#!/usr/bin/env python3
"""
End-to-end test for the Technical Competency Extraction Agent System.

This script tests the complete workflow with minimal sample data to verify
all components are working together correctly.
"""

import sys
import tempfile
import shutil
from pathlib import Path
import openpyxl

def print_status(message, status="INFO"):
    """Print colored status message."""
    colors = {
        "INFO": "\033[94m",
        "SUCCESS": "\033[92m",
        "ERROR": "\033[91m",
        "WARNING": "\033[93m"
    }
    reset = "\033[0m"
    color = colors.get(status, "")
    symbol = {
        "INFO": "ℹ",
        "SUCCESS": "✓",
        "ERROR": "✗",
        "WARNING": "⚠"
    }.get(status, "•")

    print(f"{color}{symbol} {message}{reset}")

def create_minimal_test_data(temp_dir):
    """Create minimal test data files."""
    print_status("Creating minimal test data...", "INFO")

    # Create jobs file
    wb_jobs = openpyxl.Workbook()
    ws_jobs = wb_jobs.active
    ws_jobs.title = "Jobs"
    ws_jobs.append(["Job Title", "Job Family", "Job Level", "Summary", "Responsibilities"])
    ws_jobs.append([
        "Data Scientist",
        "Analytics",
        "Senior",
        "Develop machine learning models and analyze data.",
        "Develop ML models\nAnalyze data\nPresent findings"
    ])

    jobs_file = temp_dir / "test_jobs.xlsx"
    wb_jobs.save(jobs_file)
    print_status(f"Created {jobs_file.name}", "SUCCESS")

    # Create tech competencies file
    wb_tech = openpyxl.Workbook()
    ws_tech = wb_tech.active
    ws_tech.title = "Technical"
    ws_tech.append(["Competency Name", "Definition", "Indicators", "Tags"])
    ws_tech.append([
        "Machine Learning",
        "Ability to develop and deploy machine learning models.",
        "Develops models\nEvaluates performance",
        "ML,AI,data"
    ])
    ws_tech.append([
        "Data Analysis",
        "Ability to analyze data and extract insights.",
        "Analyzes data\nCreates visualizations",
        "data,analytics"
    ])

    tech_file = temp_dir / "test_tech.xlsx"
    wb_tech.save(tech_file)
    print_status(f"Created {tech_file.name}", "SUCCESS")

    # Create leadership file
    wb_lead = openpyxl.Workbook()
    ws_lead = wb_lead.active
    ws_lead.title = "Leadership"
    ws_lead.append(["Competency Name", "Definition", "Indicators", "Tags"])
    ws_lead.append([
        "Communication",
        "Ability to communicate effectively.",
        "Presents clearly\nListens actively",
        "communication"
    ])

    lead_file = temp_dir / "test_leadership.xlsx"
    wb_lead.save(lead_file)
    print_status(f"Created {lead_file.name}", "SUCCESS")

    # Create template file
    wb_template = openpyxl.Workbook()
    ws_template = wb_template.active
    ws_template.title = "Output"
    ws_template.append([
        "Job ID", "Job Title", "Rank", "Competency Name",
        "Definition", "Why It Matters", "Behavioral Indicators",
        "Tools/Methods", "Criticality Score", "Coverage %"
    ])

    template_file = temp_dir / "test_template.xlsx"
    wb_template.save(template_file)
    print_status(f"Created {template_file.name}", "SUCCESS")

    return jobs_file, tech_file, lead_file, template_file

def test_file_analyzer():
    """Test file analyzer functionality."""
    print_status("\n=== Testing File Analyzer ===", "INFO")

    try:
        from src.utils.file_analyzer import FileAnalyzer

        # Create a simple test file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Job Title", "Summary"])
        ws.append(["Data Scientist", "Develops models"])

        test_file = Path("/tmp/test_analyzer.xlsx")
        wb.save(test_file)

        analyzer = FileAnalyzer()
        analysis = analyzer.analyze_file(test_file)

        assert analysis.file_type == "excel"
        assert analysis.row_count == 1
        assert analysis.column_count == 2

        test_file.unlink()  # Clean up

        print_status("File analyzer test passed", "SUCCESS")
        return True

    except Exception as e:
        print_status(f"File analyzer test failed: {e}", "ERROR")
        return False

def test_file_parsing():
    """Test file parsing functionality."""
    print_status("\n=== Testing File Parsing ===", "INFO")

    try:
        from src.utils.file_parsers import parse_excel_jobs, parse_competency_library

        # Create test files
        wb_jobs = openpyxl.Workbook()
        ws_jobs = wb_jobs.active
        ws_jobs.append(["Job Title", "Summary", "Responsibilities"])
        ws_jobs.append(["Test Job", "Test summary", "Task 1\nTask 2"])

        jobs_file = Path("/tmp/test_jobs.xlsx")
        wb_jobs.save(jobs_file)

        jobs, warnings = parse_excel_jobs(jobs_file)
        assert len(jobs) == 1
        assert jobs[0].job_title == "Test Job"
        assert len(jobs[0].responsibilities) == 2

        jobs_file.unlink()

        print_status("File parsing test passed", "SUCCESS")
        return True

    except Exception as e:
        print_status(f"File parsing test failed: {e}", "ERROR")
        return False

def test_similarity():
    """Test similarity computation."""
    print_status("\n=== Testing Similarity Engine ===", "INFO")

    try:
        from src.utils.similarity import compute_similarity

        score = compute_similarity(
            "Develop machine learning models",
            "Build ML models for prediction"
        )

        assert 0.0 <= score <= 1.0
        assert score > 0.5  # Should be similar

        print_status(f"Similarity test passed (score: {score:.2f})", "SUCCESS")
        return True

    except Exception as e:
        print_status(f"Similarity test failed: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False

def test_knowledge_base():
    """Test knowledge base functionality."""
    print_status("\n=== Testing Knowledge Base ===", "INFO")

    try:
        from src.utils.knowledge_base import KnowledgeBase
        import tempfile

        with tempfile.TemporaryDirectory() as temp_dir:
            kb = KnowledgeBase(Path(temp_dir))

            # Create a test document
            test_doc = Path(temp_dir) / "test.txt"
            test_doc.write_text("Machine learning is a type of artificial intelligence.")

            # Add to KB
            doc_id = kb.add_document(
                test_doc,
                title="Test Document",
                category="reference"
            )

            assert doc_id in kb.index

            # Search
            results = kb.search_documents("machine learning", top_k=1)
            assert len(results) > 0

            print_status("Knowledge base test passed", "SUCCESS")
            return True

    except Exception as e:
        print_status(f"Knowledge base test failed: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False

def test_schemas():
    """Test Pydantic schemas."""
    print_status("\n=== Testing Schemas ===", "INFO")

    try:
        from src.schemas.job import Job, Responsibility, JobSummary, SourceMetadata
        from src.schemas.competency import TechnicalCompetency, AppliedScope
        from src.schemas.run_state import RunState, RunInputs, RunConfig
        from datetime import datetime
        from pathlib import Path

        # Test Job schema
        job = Job(
            job_id="TEST_001",
            job_title="Test Job",
            job_summary=JobSummary(raw_text="Test", normalized_text="Test"),
            responsibilities=[
                Responsibility(
                    responsibility_id="TEST_001_R001",
                    raw_text="Test task",
                    normalized_text="Test task"
                )
            ],
            source_metadata=SourceMetadata()
        )

        assert job.job_id == "TEST_001"
        assert len(job.responsibilities) == 1

        # Test RunState schema
        state = RunState(
            run_id="test_run",
            inputs=RunInputs(
                jobs_file=Path("/tmp/test.xlsx"),
                tech_comp_source_files=[Path("/tmp/tech.xlsx")],
                core_leadership_file=Path("/tmp/lead.xlsx"),
                output_template_file=Path("/tmp/template.xlsx")
            ),
            config=RunConfig()
        )

        assert state.run_id == "test_run"

        print_status("Schema tests passed", "SUCCESS")
        return True

    except Exception as e:
        print_status(f"Schema tests failed: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False

def test_job_ingestion():
    """Test job ingestion agent."""
    print_status("\n=== Testing Job Ingestion Agent ===", "INFO")

    try:
        from src.agents.job_ingestion import JobIngestionAgent
        from src.schemas.run_state import RunState, RunInputs, RunConfig
        from pathlib import Path
        import tempfile

        # Create minimal test file
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Job Title", "Summary", "Responsibilities"])
            ws.append(["Test Job", "Summary", "Task 1\nTask 2\nTask 3\nTask 4\nTask 5"])

            jobs_file = temp_path / "jobs.xlsx"
            wb.save(jobs_file)

            # Create state
            state = RunState(
                run_id="test_ingestion",
                inputs=RunInputs(
                    jobs_file=jobs_file,
                    tech_comp_source_files=[temp_path / "tech.xlsx"],
                    core_leadership_file=temp_path / "lead.xlsx",
                    output_template_file=temp_path / "template.xlsx"
                ),
                config=RunConfig()
            )

            # Run agent
            agent = JobIngestionAgent("S1", "Test Ingestion")
            updated_state = agent.execute(state)

            assert updated_state.artifacts.jobs_extracted is not None
            assert updated_state.artifacts.jobs_extracted.exists()

            print_status("Job ingestion agent test passed", "SUCCESS")
            return True

    except Exception as e:
        print_status(f"Job ingestion agent test failed: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False

def run_all_tests():
    """Run all unit tests."""
    print_status("\n" + "="*60, "INFO")
    print_status("TECHNICAL COMPETENCY EXTRACTION AGENT SYSTEM", "INFO")
    print_status("End-to-End Test Suite", "INFO")
    print_status("="*60, "INFO")

    tests = [
        ("Schemas", test_schemas),
        ("File Parsing", test_file_parsing),
        ("File Analyzer", test_file_analyzer),
        ("Similarity Engine", test_similarity),
        ("Knowledge Base", test_knowledge_base),
        ("Job Ingestion Agent", test_job_ingestion),
    ]

    results = {}
    for test_name, test_func in tests:
        try:
            results[test_name] = test_func()
        except Exception as e:
            print_status(f"Unexpected error in {test_name}: {e}", "ERROR")
            results[test_name] = False

    # Summary
    print_status("\n" + "="*60, "INFO")
    print_status("TEST SUMMARY", "INFO")
    print_status("="*60, "INFO")

    passed = sum(results.values())
    total = len(results)

    for test_name, result in results.items():
        status = "SUCCESS" if result else "ERROR"
        print_status(f"{test_name}: {'PASSED' if result else 'FAILED'}", status)

    print_status(f"\nTotal: {passed}/{total} tests passed", "SUCCESS" if passed == total else "ERROR")

    if passed == total:
        print_status("\n✅ All tests passed! System is operational.", "SUCCESS")
        return 0
    else:
        print_status(f"\n⚠️  {total - passed} test(s) failed. Please review errors above.", "WARNING")
        return 1

if __name__ == "__main__":
    sys.exit(run_all_tests())
