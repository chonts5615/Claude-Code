"""End-to-end simulation: Master Skills Library + rolling Crosswalk (spec §28-29).

Verifies:
- First map-skills run creates both rolling masters with deterministic Skill_IDs.
- Idempotent re-run does NOT append duplicate crosswalk rows or version bumps.
- Material change (Bloom level shift) bumps version and appends run_history entry.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.skill_library.crosswalk_merger import (
    CROSSWALK_COLUMNS,
    merge_crosswalk,
)
from src.skill_library.skill_library import (
    SKILLS_LIBRARY_COLUMNS,
    derive_skill_id,
    merge_skills_into_master,
)
from src.skill_mapping.bloom_classifier import classify
from src.skill_mapping.schemas import BloomLevelEstimate


def _rows(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    header = list(raw[0])
    return [dict(zip(header, r)) for r in raw[1:]]


def test_first_run_creates_skills_library_and_journal(synthetic_catalog, tmp_path: Path):
    master = tmp_path / "data" / "library" / "Skills_Library_Master.xlsx"

    # Classify each catalog item so the merger can populate Bloom_* columns.
    bloom_estimates = {item.course_id: classify(item) for item in synthetic_catalog}
    merge_skills_into_master(
        synthetic_catalog,
        master,
        run_id="MAP-FIN-01",
        family="Finance",
        bloom_estimates=bloom_estimates,
    )

    assert master.exists()
    rows = _rows(master)
    assert len(rows) == len(synthetic_catalog)
    assert set(rows[0].keys()) == set(SKILLS_LIBRARY_COLUMNS)
    # IDs are deterministic.
    expected_ids = {derive_skill_id("Finance", item.course_id) for item in synthetic_catalog}
    assert {r["Skill_ID"] for r in rows} == expected_ids

    # Journal exists and has one ADD per item + one MERGE_SUMMARY.
    journal_path = master.with_suffix(".jsonl")
    events = [json.loads(line) for line in journal_path.read_text().splitlines() if line.strip()]
    adds = [e for e in events if e["event"] == "ADD"]
    summaries = [e for e in events if e["event"] == "MERGE_SUMMARY"]
    assert len(adds) == len(synthetic_catalog)
    assert len(summaries) == 1
    assert summaries[0]["adds"] == len(synthetic_catalog)
    assert summaries[0]["updates"] == 0


def test_idempotent_remerge_does_not_bump_versions(synthetic_catalog, tmp_path: Path):
    master = tmp_path / "data" / "library" / "Skills_Library_Master.xlsx"
    bloom = {item.course_id: classify(item) for item in synthetic_catalog}

    merge_skills_into_master(synthetic_catalog, master, run_id="MAP-FIN-01",
                             family="Finance", bloom_estimates=bloom)
    rows_v1 = _rows(master)

    # Re-merge identical input.
    merge_skills_into_master(synthetic_catalog, master, run_id="MAP-FIN-02",
                             family="Finance", bloom_estimates=bloom)
    rows_v2 = _rows(master)

    versions_v1 = {r["Skill_ID"]: r["Version"] for r in rows_v1}
    versions_v2 = {r["Skill_ID"]: r["Version"] for r in rows_v2}
    assert versions_v1 == versions_v2, "identical re-merge must not bump Version"

    journal_path = master.with_suffix(".jsonl")
    events = [json.loads(line) for line in journal_path.read_text().splitlines() if line.strip()]
    updates = [e for e in events if e["event"] == "UPDATE"]
    assert updates == [], "no UPDATE events on idempotent re-merge"


def test_bloom_level_shift_bumps_version(synthetic_catalog, tmp_path: Path):
    """Force a Bloom-level change on one course and confirm version bumps + UPDATE event."""
    master = tmp_path / "data" / "library" / "Skills_Library_Master.xlsx"
    initial = {item.course_id: classify(item) for item in synthetic_catalog}
    merge_skills_into_master(synthetic_catalog, master, run_id="r1",
                             family="Finance", bloom_estimates=initial)

    # Simulate a corrected Bloom level on course 1001 (L1 -> L2).
    from src.schemas.competency import LevelCode
    shifted = dict(initial)
    shifted["LDN-FIN-1001"] = BloomLevelEstimate(
        level=LevelCode.L2,
        confidence=0.78,
        evidence_verbs=["apply", "implement"],
        verb_counts={"L1": 1, "L2": 3, "L3": 0, "L4": 0},
        adjustments_applied=["AUDIENCE_BAND_SNAP"],
    )

    merge_skills_into_master(synthetic_catalog, master, run_id="r2",
                             family="Finance", bloom_estimates=shifted)

    rows = {r["Skill_ID"]: r for r in _rows(master)}
    target_sid = derive_skill_id("Finance", "LDN-FIN-1001")
    assert rows[target_sid]["Version"] == 2
    assert rows[target_sid]["Bloom_Level"] == "L2"
    assert rows[target_sid]["Last_Seen_Run"] == "r2"

    journal_path = master.with_suffix(".jsonl")
    events = [json.loads(line) for line in journal_path.read_text().splitlines() if line.strip()]
    updates = [e for e in events if e["event"] == "UPDATE" and e["skill_id"] == target_sid]
    assert len(updates) == 1
    assert "Bloom_Level" in updates[0]["changed_fields"]


def test_crosswalk_master_is_idempotent(synthetic_mappings, tmp_path: Path):
    master = tmp_path / "data" / "library" / "Skill_Competency_Crosswalk.xlsx"
    competency_names = {
        "TC-FIN-001": "Financial Risk Modeling",
        "TC-FIN-002": "Regulatory Reporting Compliance",
        "TC-FIN-003": "Treasury Cash Forecasting",
    }

    merge_crosswalk(synthetic_mappings, master, run_id="r1",
                    competency_names=competency_names)
    rows_v1 = _rows(master)
    assert len(rows_v1) == len(synthetic_mappings)
    assert set(rows_v1[0].keys()) == set(CROSSWALK_COLUMNS)
    histories_v1 = {(r["Course_ID"], r["Competency_ID"]): json.loads(r["Run_History"]) for r in rows_v1}
    assert all(len(h) == 1 for h in histories_v1.values())

    # Re-run identical mappings; the master must not grow and run_history must not grow.
    merge_crosswalk(synthetic_mappings, master, run_id="r2",
                    competency_names=competency_names)
    rows_v2 = _rows(master)
    assert len(rows_v2) == len(synthetic_mappings)
    histories_v2 = {(r["Course_ID"], r["Competency_ID"]): json.loads(r["Run_History"]) for r in rows_v2}
    for key, h in histories_v2.items():
        assert len(h) == 1, f"identical re-run must not append run_history for {key}; got {h}"
    # Last_Mapped_Run should still bump to r2.
    assert all(r["Last_Mapped_Run"] == "r2" for r in rows_v2)


def test_crosswalk_logs_material_change(synthetic_mappings, tmp_path: Path):
    """Changing one mapping's confidence appends a run_history entry but does not duplicate the row."""
    master = tmp_path / "data" / "library" / "Skill_Competency_Crosswalk.xlsx"
    merge_crosswalk(synthetic_mappings, master, run_id="r1")

    # Bump confidence on the first mapping.
    mutated = list(synthetic_mappings)
    mutated[0] = synthetic_mappings[0].model_copy(update={
        "confidence": 0.85,
        "integrity_tag": "CONFIRMED",
    })

    merge_crosswalk(mutated, master, run_id="r2")
    rows = _rows(master)
    assert len(rows) == len(synthetic_mappings), "must not duplicate rows on material change"

    target = next(r for r in rows if r["Course_ID"] == "LDN-FIN-1001"
                  and r["Competency_ID"] == "TC-FIN-001")
    assert float(target["Current_Confidence"]) == pytest.approx(0.85)
    history = json.loads(target["Run_History"])
    assert len(history) == 2, f"material change must append history; got {history}"
    assert history[-1]["confidence"] == pytest.approx(0.85)
    assert history[-1]["run_id"] == "r2"
