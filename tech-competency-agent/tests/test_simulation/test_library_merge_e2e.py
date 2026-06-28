"""End-to-end simulation: Master Library cross-run merger (spec §30.5).

Verifies:
- First run creates the rolling master with all entries.
- Second identical run is idempotent (zero change_log events).
- Second run with one edited definition emits exactly one change_log event for
  that field, and the row is updated in place.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.schemas.library import LIBRARY_COLUMNS
from src.skill_library.library_merger import merge_library_into_master
from src.tracing.ledger import read_all


@pytest.fixture
def trace_root(tmp_path: Path, monkeypatch) -> Path:
    """Redirect tracing AND the master library output to tmp_path."""
    monkeypatch.chdir(tmp_path)
    return tmp_path / "data" / "trace"


def _read_master(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    return [dict(zip(header, r)) for r in rows[1:]]


def test_first_merge_creates_master_and_logs_creation(
    trace_root: Path, synthetic_competencies, tmp_path: Path
):
    master = tmp_path / "data" / "library" / "TechComp_Library_Master.xlsx"
    merge_library_into_master(
        synthetic_competencies, master, run_id="Finance_R1_01", family="Finance"
    )

    assert master.exists()
    rows = _read_master(master)
    assert len(rows) == 3
    assert {r["competency_id"] for r in rows} == {"TC-FIN-001", "TC-FIN-002", "TC-FIN-003"}
    # All marked with the run id.
    assert all(r["first_published_run"] == "Finance_R1_01" for r in rows)
    # Columns match canonical schema.
    assert set(rows[0].keys()) == set(LIBRARY_COLUMNS)

    changes = read_all("change_log.jsonl", root=trace_root)
    creates = [c for c in changes if c["field"] == "<created>"]
    assert len(creates) == 3, "one creation event per new competency"


def test_idempotent_remerge_emits_no_change_events(
    trace_root: Path, synthetic_competencies, tmp_path: Path
):
    master = tmp_path / "data" / "library" / "TechComp_Library_Master.xlsx"
    merge_library_into_master(synthetic_competencies, master, run_id="r1", family="Finance")
    changes_after_first = len(read_all("change_log.jsonl", root=trace_root))

    # Re-merge identical input.
    merge_library_into_master(synthetic_competencies, master, run_id="r1", family="Finance")
    changes_after_second = len(read_all("change_log.jsonl", root=trace_root))

    assert changes_after_second == changes_after_first, (
        "idempotent re-merge must produce zero new change events; "
        f"first={changes_after_first} second={changes_after_second}"
    )


def test_definition_edit_emits_single_field_change(
    trace_root: Path, synthetic_competencies, tmp_path: Path
):
    master = tmp_path / "data" / "library" / "TechComp_Library_Master.xlsx"
    merge_library_into_master(synthetic_competencies, master, run_id="r1", family="Finance")

    # Edit one competency's definition; everything else identical.
    edited = list(synthetic_competencies)
    edited[0] = synthetic_competencies[0].model_copy(update={
        "definition": "Builds quantitative models forecasting credit and market risk across "
                      "all trading and portfolio decisions every single trading day.",
    })

    pre_changes = len(read_all("change_log.jsonl", root=trace_root))
    merge_library_into_master(edited, master, run_id="r2", family="Finance")
    post_changes = read_all("change_log.jsonl", root=trace_root)

    new_changes = post_changes[pre_changes:]
    definition_changes = [c for c in new_changes if c["field"] == "definition"]
    assert len(definition_changes) == 1, f"expected one definition change; got {new_changes}"
    assert definition_changes[0]["competency_id"] == "TC-FIN-001"
    assert "Builds quantitative models forecasting" in definition_changes[0]["after"]

    # Row updated in place; first_published_run preserved.
    rows = {r["competency_id"]: r for r in _read_master(master)}
    assert "all trading and portfolio decisions" in str(rows["TC-FIN-001"]["definition"])
    assert rows["TC-FIN-001"]["first_published_run"] == "r1"
    assert rows["TC-FIN-001"]["last_modified_run"] == "r2"
