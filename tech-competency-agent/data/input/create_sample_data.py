"""Create sample input files for testing."""

import openpyxl
from pathlib import Path

# Create sample jobs Excel file
wb_jobs = openpyxl.Workbook()
ws_jobs = wb_jobs.active
ws_jobs.title = "Jobs"

# Headers
ws_jobs.append(["Job Title", "Job Family", "Job Level", "Summary", "Responsibilities"])

# Sample jobs
ws_jobs.append([
    "Senior Data Scientist",
    "Data & Analytics",
    "Senior",
    "Lead complex data science projects, develop machine learning models, and provide strategic insights to drive business decisions.",
    """Develop and deploy machine learning models for predictive analytics
Design and implement data pipelines for large-scale data processing
Collaborate with cross-functional teams to translate business requirements into analytical solutions
Conduct exploratory data analysis to uncover insights and patterns
Mentor junior data scientists and provide technical guidance
Present findings and recommendations to senior stakeholders
Optimize model performance through feature engineering and hyperparameter tuning"""
])

ws_jobs.append([
    "Software Engineer",
    "Engineering",
    "Mid-Level",
    "Design, develop, and maintain software applications using modern technologies and best practices.",
    """Write clean, efficient, and maintainable code
Participate in code reviews and provide constructive feedback
Collaborate with product managers to define feature requirements
Debug and resolve software defects and performance issues
Contribute to architectural design discussions
Write unit and integration tests to ensure code quality
Deploy applications to production environments
Document technical specifications and processes"""
])

ws_jobs.append([
    "Cloud Solutions Architect",
    "Infrastructure",
    "Senior",
    "Design and implement scalable cloud infrastructure solutions, ensuring security, reliability, and cost-effectiveness.",
    """Design cloud architecture solutions using AWS, Azure, or GCP
Develop infrastructure-as-code using Terraform or CloudFormation
Implement security best practices and compliance requirements
Optimize cloud costs and resource utilization
Lead migration of on-premise applications to cloud
Establish disaster recovery and business continuity plans
Mentor development teams on cloud-native patterns"""
])

wb_jobs.save("sample_jobs.xlsx")
print("✓ Created sample_jobs.xlsx")

# Create sample technical competencies library
wb_comps = openpyxl.Workbook()
ws_comps = wb_comps.active
ws_comps.title = "Technical Competencies"

# Headers
ws_comps.append(["Competency Name", "Definition", "Indicators", "Tags"])

# Sample competencies
ws_comps.append([
    "Machine Learning: Model Development",
    "Ability to design, develop, and deploy machine learning models to solve business problems. Includes selecting appropriate algorithms, feature engineering, model training, and evaluation.",
    """Selects appropriate ML algorithms based on problem type and data characteristics
Performs feature engineering to improve model performance
Trains models using cross-validation and hyperparameter tuning
Evaluates model performance using appropriate metrics
Deploys models to production environments
Monitors model performance and implements retraining strategies""",
    "machine learning, ML, AI, data science"
])

ws_comps.append([
    "Data Engineering: Pipeline Development",
    "Ability to design and implement data pipelines that efficiently process, transform, and move data across systems. Includes ETL/ELT processes, data quality, and orchestration.",
    """Designs scalable data pipeline architectures
Implements ETL/ELT processes using tools like Airflow, Spark, or dbt
Ensures data quality through validation and monitoring
Optimizes pipeline performance for large-scale data
Handles data schema evolution and versioning
Implements error handling and retry mechanisms""",
    "data engineering, ETL, pipelines, data processing"
])

ws_comps.append([
    "Software Development: Application Design",
    "Ability to design software applications with clean architecture, following design patterns and best practices for maintainability and scalability.",
    """Applies SOLID principles and design patterns
Creates modular, reusable code components
Designs RESTful APIs following industry standards
Implements proper error handling and logging
Writes comprehensive unit and integration tests
Documents code and architectural decisions""",
    "software development, programming, design patterns"
])

ws_comps.append([
    "Cloud Architecture: Infrastructure Design",
    "Ability to design and implement cloud infrastructure that is scalable, secure, and cost-effective using major cloud platforms.",
    """Designs multi-tier cloud architectures
Implements infrastructure-as-code using Terraform or CloudFormation
Applies cloud security best practices
Optimizes for cost and performance
Implements auto-scaling and load balancing
Designs for high availability and disaster recovery""",
    "cloud, AWS, Azure, GCP, infrastructure"
])

ws_comps.append([
    "Programming: Python Development",
    "Proficiency in Python programming for developing applications, scripts, and data analysis tools following Python best practices.",
    """Writes idiomatic Python code following PEP 8 standards
Uses Python libraries effectively (pandas, numpy, scikit-learn)
Implements object-oriented and functional programming patterns
Writes unit tests using pytest or unittest
Manages dependencies using pip, poetry, or conda
Debugs and profiles Python applications""",
    "Python, programming, coding"
])

wb_comps.save("sample_tech_competencies.xlsx")
print("✓ Created sample_tech_competencies.xlsx")

# Create sample leadership competencies
wb_lead = openpyxl.Workbook()
ws_lead = wb_lead.active
ws_lead.title = "Leadership Competencies"

ws_lead.append(["Competency Name", "Definition", "Indicators", "Tags"])

ws_lead.append([
    "Leadership: Team Development",
    "Ability to develop team members through coaching, mentoring, and providing growth opportunities.",
    """Provides regular feedback and coaching to team members
Identifies development needs and creates growth plans
Delegates challenging assignments to build skills
Recognizes and rewards team achievements
Creates inclusive team culture""",
    "leadership, management, coaching"
])

ws_lead.append([
    "Communication: Stakeholder Management",
    "Ability to effectively communicate with stakeholders at all levels, tailoring message to audience.",
    """Presents complex information clearly to non-technical audiences
Actively listens and incorporates stakeholder feedback
Manages expectations and negotiates trade-offs
Builds relationships across organizational boundaries
Facilitates productive meetings and discussions""",
    "communication, stakeholder, presentation"
])

ws_lead.append([
    "Strategic Thinking: Business Alignment",
    "Ability to align technical decisions with business strategy and organizational goals.",
    """Understands business context and market dynamics
Translates business requirements into technical solutions
Identifies opportunities for strategic impact
Balances short-term needs with long-term vision
Makes data-driven decisions""",
    "strategy, business, alignment"
])

wb_lead.save("sample_leadership.xlsx")
print("✓ Created sample_leadership.xlsx")

# Create sample output template
wb_template = openpyxl.Workbook()
ws_template = wb_template.active
ws_template.title = "Technical Competencies"

# Headers with formatting
headers = ["Job ID", "Job Title", "Rank", "Competency Name", "Definition",
           "Why It Matters", "Behavioral Indicators", "Tools/Methods/Technologies",
           "Criticality Score", "Responsibility Coverage %"]
ws_template.append(headers)

# Make headers bold
for cell in ws_template[1]:
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

# Set column widths
ws_template.column_dimensions['A'].width = 15
ws_template.column_dimensions['B'].width = 30
ws_template.column_dimensions['C'].width = 8
ws_template.column_dimensions['D'].width = 40
ws_template.column_dimensions['E'].width = 60
ws_template.column_dimensions['F'].width = 50
ws_template.column_dimensions['G'].width = 60
ws_template.column_dimensions['H'].width = 40
ws_template.column_dimensions['I'].width = 15
ws_template.column_dimensions['J'].width = 20

wb_template.save("sample_template.xlsx")
print("✓ Created sample_template.xlsx")

print("\n=== Sample files created successfully! ===")
print("\nTo test the workflow, run:")
print("  cd /home/user/Claude-Code/tech-competency-agent")
print("  python data/input/create_sample_data.py")
print("  techcomp run \\")
print("    --jobs-file data/input/sample_jobs.xlsx \\")
print("    --tech-sources data/input/sample_tech_competencies.xlsx \\")
print("    --leadership-file data/input/sample_leadership.xlsx \\")
print("    --template-file data/input/sample_template.xlsx")
