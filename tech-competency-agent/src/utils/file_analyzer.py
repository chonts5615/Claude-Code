"""File analysis utilities for inspecting and understanding input files."""

from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
import pandas as pd
from pydantic import BaseModel


class ColumnInfo(BaseModel):
    """Information about a column in a file."""
    name: str
    index: int
    sample_values: List[str]
    non_null_count: int
    total_rows: int
    data_type: str
    suggested_purpose: Optional[str] = None


class FileAnalysis(BaseModel):
    """Analysis results for a file."""
    file_path: Path
    file_type: str
    sheet_names: List[str] = []
    active_sheet: Optional[str] = None
    row_count: int = 0
    column_count: int = 0
    columns: List[ColumnInfo] = []
    suggested_file_purpose: Optional[str] = None
    confidence_score: float = 0.0


class FileAnalyzer:
    """Analyzes input files to understand their structure and purpose."""

    def analyze_file(self, file_path: Path) -> FileAnalysis:
        """
        Analyze a file and return structural information.

        Args:
            file_path: Path to file to analyze

        Returns:
            FileAnalysis with detected structure
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = file_path.suffix.lower()

        if suffix in ['.xlsx', '.xls']:
            return self._analyze_excel(file_path)
        elif suffix == '.csv':
            return self._analyze_csv(file_path)
        else:
            return FileAnalysis(
                file_path=file_path,
                file_type=suffix,
                suggested_file_purpose="UNKNOWN"
            )

    def _analyze_excel(self, file_path: Path) -> FileAnalysis:
        """Analyze Excel file structure."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        active_sheet = wb.active.title

        # Analyze active sheet
        sheet = wb.active

        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Analyze columns
        columns = []
        headers = []

        # Try to detect headers (assume first row)
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers.append((col_idx, str(cell.value).strip()))

        # Get sample data for each column
        for col_idx, col_name in headers:
            sample_values = []
            non_null = 0

            for row_idx in range(2, min(12, max_row + 1)):  # Sample first 10 data rows
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    sample_values.append(str(cell.value).strip()[:100])  # Truncate long values
                    non_null += 1

            # Detect data type
            data_type = self._infer_data_type(sample_values)

            # Suggest purpose based on column name and content
            suggested_purpose = self._suggest_column_purpose(col_name, sample_values, data_type)

            columns.append(ColumnInfo(
                name=col_name,
                index=col_idx - 1,
                sample_values=sample_values[:5],
                non_null_count=non_null,
                total_rows=max_row - 1,
                data_type=data_type,
                suggested_purpose=suggested_purpose
            ))

        wb.close()

        # Suggest file purpose
        file_purpose, confidence = self._suggest_file_purpose(columns)

        return FileAnalysis(
            file_path=file_path,
            file_type="excel",
            sheet_names=sheet_names,
            active_sheet=active_sheet,
            row_count=max_row - 1,  # Exclude header
            column_count=len(columns),
            columns=columns,
            suggested_file_purpose=file_purpose,
            confidence_score=confidence
        )

    def _analyze_csv(self, file_path: Path) -> FileAnalysis:
        """Analyze CSV file structure."""
        df = pd.read_csv(file_path, nrows=10)  # Sample first 10 rows

        columns = []
        for col_name in df.columns:
            sample_values = df[col_name].dropna().astype(str).head(5).tolist()
            non_null = df[col_name].notna().sum()
            data_type = self._infer_data_type(sample_values)
            suggested_purpose = self._suggest_column_purpose(str(col_name), sample_values, data_type)

            columns.append(ColumnInfo(
                name=str(col_name),
                index=df.columns.get_loc(col_name),
                sample_values=sample_values,
                non_null_count=int(non_null),
                total_rows=len(df),
                data_type=data_type,
                suggested_purpose=suggested_purpose
            ))

        file_purpose, confidence = self._suggest_file_purpose(columns)

        return FileAnalysis(
            file_path=file_path,
            file_type="csv",
            row_count=len(df),
            column_count=len(columns),
            columns=columns,
            suggested_file_purpose=file_purpose,
            confidence_score=confidence
        )

    def _infer_data_type(self, sample_values: List[str]) -> str:
        """Infer data type from sample values."""
        if not sample_values:
            return "unknown"

        # Check if numeric
        try:
            [float(v) for v in sample_values if v]
            return "numeric"
        except ValueError:
            pass

        # Check average length
        avg_length = sum(len(v) for v in sample_values) / len(sample_values)

        if avg_length > 200:
            return "long_text"
        elif avg_length > 50:
            return "text"
        else:
            return "short_text"

    def _suggest_column_purpose(self, col_name: str, sample_values: List[str], data_type: str) -> str:
        """Suggest the purpose of a column based on name and content."""
        col_name_lower = col_name.lower()

        # Job-related columns
        if any(term in col_name_lower for term in ['job title', 'title', 'position', 'role']):
            return "job_title"
        elif any(term in col_name_lower for term in ['job family', 'family', 'department']):
            return "job_family"
        elif any(term in col_name_lower for term in ['level', 'grade', 'tier']):
            return "job_level"
        elif any(term in col_name_lower for term in ['summary', 'description', 'overview']):
            return "job_summary"
        elif any(term in col_name_lower for term in ['responsibilities', 'duties', 'tasks']):
            return "responsibilities"
        elif any(term in col_name_lower for term in ['id', 'code', 'number']) and data_type in ['short_text', 'numeric']:
            return "identifier"

        # Competency-related columns
        elif any(term in col_name_lower for term in ['competency', 'skill', 'ability']):
            if 'name' in col_name_lower or 'title' in col_name_lower:
                return "competency_name"
            elif 'definition' in col_name_lower or 'description' in col_name_lower:
                return "competency_definition"
            else:
                return "competency_field"
        elif any(term in col_name_lower for term in ['indicator', 'behavior']):
            return "behavioral_indicators"
        elif 'tag' in col_name_lower or 'category' in col_name_lower:
            return "tags"

        # Generic
        elif data_type == "long_text":
            return "text_content"
        else:
            return "general_data"

    def _suggest_file_purpose(self, columns: List[ColumnInfo]) -> Tuple[str, float]:
        """
        Suggest the overall purpose of the file based on columns.

        Returns:
            Tuple of (purpose, confidence_score)
        """
        purpose_counts = {}

        for col in columns:
            if col.suggested_purpose:
                category = self._categorize_purpose(col.suggested_purpose)
                purpose_counts[category] = purpose_counts.get(category, 0) + 1

        if not purpose_counts:
            return "UNKNOWN", 0.0

        # Determine primary purpose
        max_category = max(purpose_counts, key=purpose_counts.get)
        confidence = purpose_counts[max_category] / len(columns)

        # Map to file types
        if max_category == "job":
            if purpose_counts[max_category] >= 3:  # Need at least 3 job-related columns
                return "JOBS_FILE", min(confidence * 1.5, 1.0)
            else:
                return "POSSIBLY_JOBS_FILE", confidence
        elif max_category == "competency":
            if purpose_counts[max_category] >= 2:
                return "COMPETENCY_LIBRARY", min(confidence * 1.5, 1.0)
            else:
                return "POSSIBLY_COMPETENCY_LIBRARY", confidence
        else:
            return "GENERAL_DATA", confidence

    def _categorize_purpose(self, purpose: str) -> str:
        """Categorize a column purpose into broader categories."""
        job_purposes = ['job_title', 'job_family', 'job_level', 'job_summary', 'responsibilities']
        competency_purposes = ['competency_name', 'competency_definition', 'competency_field',
                              'behavioral_indicators', 'tags']

        if purpose in job_purposes:
            return "job"
        elif purpose in competency_purposes:
            return "competency"
        else:
            return "other"

    def suggest_column_mapping(self, analysis: FileAnalysis, target_type: str) -> Dict[str, str]:
        """
        Suggest column mapping for a file based on target type.

        Args:
            analysis: File analysis results
            target_type: Target type ('jobs', 'competencies', 'leadership')

        Returns:
            Dictionary mapping target fields to column names
        """
        mapping = {}

        if target_type == "jobs":
            # Find best match for each required field
            for col in analysis.columns:
                if col.suggested_purpose == "job_title":
                    mapping["Job Title"] = col.name
                elif col.suggested_purpose == "job_family":
                    mapping["Job Family"] = col.name
                elif col.suggested_purpose == "job_level":
                    mapping["Job Level"] = col.name
                elif col.suggested_purpose == "job_summary":
                    mapping["Summary"] = col.name
                elif col.suggested_purpose == "responsibilities":
                    mapping["Responsibilities"] = col.name

        elif target_type == "competencies":
            for col in analysis.columns:
                if col.suggested_purpose == "competency_name":
                    mapping["Competency Name"] = col.name
                elif col.suggested_purpose == "competency_definition":
                    mapping["Definition"] = col.name
                elif col.suggested_purpose == "behavioral_indicators":
                    mapping["Indicators"] = col.name
                elif col.suggested_purpose == "tags":
                    mapping["Tags"] = col.name

        return mapping
