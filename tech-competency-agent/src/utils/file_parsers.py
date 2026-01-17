"""File parsing utilities for Excel, Word, and PDF files."""

from pathlib import Path
from typing import List, Tuple
import openpyxl
from datetime import datetime
import re

from src.schemas.job import Job, Responsibility, JobSummary, SourceMetadata, ExtractionWarning
from src.schemas.competency import CompetencyLibraryEntry, SourceEvidence


def parse_excel_jobs(file_path: Path) -> Tuple[List[Job], List[ExtractionWarning]]:
    """
    Parse jobs from Excel file.

    Expected format:
    - Each row is a job
    - Columns: Job Title, Job Family, Job Level, Summary, Responsibilities

    Args:
        file_path: Path to Excel file

    Returns:
        Tuple of (jobs list, warnings list)
    """
    jobs = []
    warnings = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Detect header row (simple implementation - assumes row 1)
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # Map column names to indices
        col_mapping = {header: idx for idx, header in enumerate(headers)}

        # Process each row
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue

            try:
                job = _parse_job_row(row, col_mapping, sheet.title, row_idx)
                if job:
                    jobs.append(job)

                    # Validate minimum responsibilities
                    if job.responsibility_count() < 5:
                        warnings.append(ExtractionWarning(
                            job_id=job.job_id,
                            warning_type="NO_RESPONSIBILITIES",
                            message=f"Job has only {job.responsibility_count()} responsibilities (minimum 5 expected)",
                            severity="WARNING"
                        ))

                    # Check for missing summary
                    if not job.job_summary.raw_text.strip():
                        warnings.append(ExtractionWarning(
                            job_id=job.job_id,
                            warning_type="MISSING_SUMMARY",
                            message=f"Job missing summary",
                            severity="WARNING"
                        ))

            except Exception as e:
                warnings.append(ExtractionWarning(
                    warning_type="OTHER",
                    message=f"Error parsing row {row_idx}: {str(e)}",
                    severity="ERROR"
                ))

        wb.close()

    except Exception as e:
        warnings.append(ExtractionWarning(
            warning_type="OTHER",
            message=f"Error reading Excel file: {str(e)}",
            severity="ERROR"
        ))

    return jobs, warnings


def _parse_job_row(
    row: tuple,
    col_mapping: dict,
    sheet_name: str,
    row_idx: int
) -> Job:
    """Parse a single job row."""
    # Extract fields
    job_title = _get_cell_value(row, col_mapping, "Job Title", "")
    if not job_title:
        return None

    job_family = _get_cell_value(row, col_mapping, "Job Family", None)
    job_level = _get_cell_value(row, col_mapping, "Job Level", None)
    summary_text = _get_cell_value(row, col_mapping, "Summary", "")
    responsibilities_text = _get_cell_value(row, col_mapping, "Responsibilities", "")

    # Create job ID
    job_id = f"JOB_{row_idx:04d}"

    # Parse responsibilities
    responsibilities = _parse_responsibilities(responsibilities_text, job_id)

    # Create job
    job = Job(
        job_id=job_id,
        job_title=job_title,
        job_family=job_family,
        job_level=job_level,
        job_summary=JobSummary(
            raw_text=summary_text,
            normalized_text=_normalize_text(summary_text)
        ),
        responsibilities=responsibilities,
        source_metadata=SourceMetadata(
            sheet_name=sheet_name,
            row_index=row_idx,
            column_mapping={k: str(v) for k, v in col_mapping.items()},
            extraction_timestamp=datetime.utcnow().isoformat()
        )
    )

    return job


def _get_cell_value(row: tuple, col_mapping: dict, col_name: str, default=None):
    """Get cell value from row by column name."""
    if col_name in col_mapping:
        idx = col_mapping[col_name]
        if idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip()
    return default


def _parse_responsibilities(text: str, job_id: str) -> List[Responsibility]:
    """
    Parse responsibilities from text.

    Supports:
    - Newline-separated
    - Bullet points (•, -, *)
    - Numbered lists (1., 2., etc.)
    """
    if not text:
        return []

    responsibilities = []

    # Split by newlines first
    lines = text.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Remove bullet points and numbering
        line = re.sub(r'^[•\-\*]\s*', '', line)
        line = re.sub(r'^\d+[\.\)]\s*', '', line)

        if line:
            resp_id = f"{job_id}_R{len(responsibilities)+1:03d}"
            responsibilities.append(Responsibility(
                responsibility_id=resp_id,
                raw_text=line,
                normalized_text=_normalize_text(line),
                priority_hint="UNKNOWN"
            ))

    return responsibilities


def _normalize_text(text: str) -> str:
    """Normalize text by cleaning whitespace and formatting."""
    if not text:
        return ""

    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()

    return text


def parse_competency_library(file_path: Path) -> List[CompetencyLibraryEntry]:
    """
    Parse competency library from Excel file.

    Expected format:
    - Each row is a competency
    - Columns: Competency Name, Definition, Indicators, Tags

    Args:
        file_path: Path to Excel file

    Returns:
        List of competency library entries
    """
    competencies = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Detect headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        col_mapping = {header: idx for idx, header in enumerate(headers)}

        # Process each row
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            comp_name = _get_cell_value(row, col_mapping, "Competency Name", "")
            if not comp_name:
                continue

            comp_id = f"COMP_{file_path.stem}_{row_idx:04d}"
            definition = _get_cell_value(row, col_mapping, "Definition", "")
            indicators_text = _get_cell_value(row, col_mapping, "Indicators", "")
            tags_text = _get_cell_value(row, col_mapping, "Tags", "")

            # Parse indicators
            indicators = []
            if indicators_text:
                indicators = [i.strip() for i in indicators_text.split('\n') if i.strip()]

            # Parse tags
            tags = []
            if tags_text:
                tags = [t.strip() for t in tags_text.split(',') if t.strip()]

            # Create source evidence
            evidence = SourceEvidence(
                source_id=comp_id,
                source_type="EXCEL",
                source_title=file_path.name,
                excerpt=definition[:200],
                location=f"Sheet: {sheet.title}, Row: {row_idx}",
                retrieval_date_utc=datetime.utcnow().isoformat()
            )

            competencies.append(CompetencyLibraryEntry(
                competency_id=comp_id,
                name=comp_name,
                definition=definition,
                indicators=indicators,
                tags=tags,
                source_evidence=[evidence]
            ))

        wb.close()

    except Exception as e:
        print(f"Error parsing competency library: {str(e)}")

    return competencies


def parse_word_document(file_path: Path):
    """Parse Word document (TODO: implement)."""
    # TODO: Implement using python-docx
    raise NotImplementedError("Word document parsing not yet implemented")


def parse_pdf_document(file_path: Path):
    """Parse PDF document (TODO: implement)."""
    # TODO: Implement using pypdf
    raise NotImplementedError("PDF parsing not yet implemented")
