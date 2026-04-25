"""Cargill brand constants for v3.1 deliverables.

Intentionally duplicated from cargill-pptx-converter/src/brand/constants.py per
the scope decision to keep this repo standalone. When a shared `cargill-brand`
package exists, replace this module with `from cargill_brand import ...`.
"""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill

LEAF_GREEN = "#00843D"
WHITE_GREEN = "#F5F9ED"
BLACK = "#000000"

ARIAL = "Arial"
GEORGIA = "Georgia"
BIG_CASLON = "Big Caslon"


def _hex(color: str) -> str:
    return color.lstrip("#").upper()


def header_fill() -> PatternFill:
    return PatternFill(start_color=_hex(LEAF_GREEN), end_color=_hex(LEAF_GREEN), fill_type="solid")


def alt_row_fill() -> PatternFill:
    return PatternFill(start_color=_hex(WHITE_GREEN), end_color=_hex(WHITE_GREEN), fill_type="solid")


def header_font() -> Font:
    return Font(name=ARIAL, bold=True, color="FFFFFF", size=11)


def body_font() -> Font:
    return Font(name=ARIAL, size=10)
