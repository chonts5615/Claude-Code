"""SM8 — Branded Skill→Competency→Level crosswalk Excel writer."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schemas.competency import LevelCode
from src.skill_mapping.schemas import (
    CoverageCell,
    GapFinding,
    SkillCompetencyMapping,
    SurplusFinding,
    TrainingItem,
)

try:  # Prefer shared brand helpers; fall back if the module is unavailable.
    from src.utils.branding import (
        ARIAL,
        LEAF_GREEN,
        WHITE_GREEN,
        alt_row_fill,
        body_font,
        header_fill,
        header_font,
    )

    _BRANDING_AVAILABLE = True
except ImportError:  # pragma: no cover — TODO: import from src.utils.branding once available
    _BRANDING_AVAILABLE = False
    ARIAL = "Arial"
    LEAF_GREEN = "#00843D"
    WHITE_GREEN = "#F5F9ED"

    def header_fill() -> PatternFill:
        return PatternFill(
            start_color="00843D", end_color="00843D", fill_type="solid"
        )

    def alt_row_fill() -> PatternFill:
        return PatternFill(
            start_color="F5F9ED", end_color="F5F9ED", fill_type="solid"
        )

    def header_font() -> Font:
        return Font(name=ARIAL, bold=True, color="FFFFFF", size=11)

    def body_font() -> Font:
        return Font(name=ARIAL, size=10)


_GAP_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
_LEVELS: list[LevelCode] = [LevelCode.L1, LevelCode.L2, LevelCode.L3, LevelCode.L4]


def _apply_header(ws: Worksheet, headers: list[str]) -> None:
    fill = header_fill()
    font = header_font()
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _apply_body_row(ws: Worksheet, row_idx: int, values: list, alt: bool) -> None:
    fill = alt_row_fill() if alt else None
    font = body_font()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if fill is not None:
            cell.fill = fill


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        widths = [len(str(c.value)) if c.value is not None else 0 for c in col_cells]
        ws.column_dimensions[col_letter].width = min(max(10, max(widths) + 2), max_width)


def _write_metadata(
    ws: Worksheet,
    family: str,
    items: list[TrainingItem],
    mappings: list[SkillCompetencyMapping],
    coverage: list[CoverageCell],
) -> None:
    total_courses = len(items)
    mapped_ids = {m.course_id for m in mappings}
    unmapped_pct = (
        100.0 * (total_courses - len(mapped_ids)) / total_courses if total_courses else 0.0
    )
    by_comp: dict[str, list[CoverageCell]] = {}
    for c in coverage:
        by_comp.setdefault(c.competency_id, []).append(c)
    zero_training = sum(
        1 for cells in by_comp.values() if sum(c.count for c in cells) == 0
    )
    zero_pct = 100.0 * zero_training / len(by_comp) if by_comp else 0.0

    rows = [
        ("Family", family),
        ("Generated (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Total courses", total_courses),
        ("Total mappings", len(mappings)),
        ("Unmapped courses (%)", round(unmapped_pct, 2)),
        ("Zero-training competencies (%)", round(zero_pct, 2)),
        ("Branding source", "src.utils.branding" if _BRANDING_AVAILABLE else "fallback (TODO)"),
    ]
    _apply_header(ws, ["Field", "Value"])
    for i, (k, v) in enumerate(rows, start=2):
        _apply_body_row(ws, i, [k, v], alt=(i % 2 == 0))
    _autosize(ws)


def _write_crosswalk_with_titles(
    ws: Worksheet,
    mappings: list[SkillCompetencyMapping],
    items: list[TrainingItem],
) -> None:
    headers = [
        "course_id",
        "course_title",
        "competency_id",
        "competency_name",
        "level",
        "confidence",
        "similarity",
        "integrity_tag",
        "mismatch_flags",
        "rationale",
    ]
    title_lookup = {i.course_id: i.title for i in items}
    _apply_header(ws, headers)
    for i, m in enumerate(mappings, start=2):
        _apply_body_row(
            ws,
            i,
            [
                m.course_id,
                title_lookup.get(m.course_id, ""),
                m.competency_id,
                m.competency_name,
                m.level.value,
                round(m.confidence, 3),
                round(m.similarity_score, 3),
                m.integrity_tag,
                "|".join(m.mismatch_flags),
                m.rationale,
            ],
            alt=(i % 2 == 0),
        )
    _autosize(ws)


def _write_coverage(ws: Worksheet, coverage: list[CoverageCell]) -> None:
    headers = ["competency_id", "competency_name"] + [lvl.value for lvl in _LEVELS]
    _apply_header(ws, headers)

    by_comp: dict[str, dict[LevelCode, CoverageCell]] = {}
    name_lookup: dict[str, str] = {}
    for cell in coverage:
        by_comp.setdefault(cell.competency_id, {})[cell.level] = cell
        name_lookup[cell.competency_id] = cell.competency_name

    for row_idx, cid in enumerate(sorted(by_comp.keys()), start=2):
        row = [cid, name_lookup.get(cid, "")]
        for lvl in _LEVELS:
            cell = by_comp[cid].get(lvl)
            row.append(cell.count if cell else 0)
        _apply_body_row(ws, row_idx, row, alt=(row_idx % 2 == 0))
        # Highlight any zero-count level cell.
        for offset, lvl in enumerate(_LEVELS):
            level_cell = ws.cell(row=row_idx, column=3 + offset)
            if level_cell.value == 0:
                level_cell.fill = _GAP_FILL
    _autosize(ws, max_width=30)


def _write_gaps(ws: Worksheet, gaps: list[GapFinding]) -> None:
    headers = ["competency_id", "competency_name", "level", "severity", "recommendation"]
    _apply_header(ws, headers)
    for i, g in enumerate(gaps, start=2):
        _apply_body_row(
            ws,
            i,
            [g.competency_id, g.competency_name, g.level.value, g.severity, g.recommendation],
            alt=(i % 2 == 0),
        )
    _autosize(ws)


def _write_surplus(ws: Worksheet, surplus: list[SurplusFinding]) -> None:
    headers = ["course_id", "title", "reason", "max_similarity"]
    _apply_header(ws, headers)
    rows = [s for s in surplus if s.reason != "LIKELY_VB_OR_COMMON"]
    for i, s in enumerate(rows, start=2):
        _apply_body_row(
            ws,
            i,
            [s.course_id, s.title, s.reason, round(s.max_similarity, 3)],
            alt=(i % 2 == 0),
        )
    _autosize(ws)


def _write_vb_common(ws: Worksheet, surplus: list[SurplusFinding]) -> None:
    headers = ["course_id", "title", "reason", "max_similarity"]
    _apply_header(ws, headers)
    rows = [s for s in surplus if s.reason == "LIKELY_VB_OR_COMMON"]
    for i, s in enumerate(rows, start=2):
        _apply_body_row(
            ws,
            i,
            [s.course_id, s.title, s.reason, round(s.max_similarity, 3)],
            alt=(i % 2 == 0),
        )
    _autosize(ws)


def write_crosswalk(
    out_path: Path,
    mappings: list[SkillCompetencyMapping],
    coverage: list[CoverageCell],
    gaps: list[GapFinding],
    surplus: list[SurplusFinding],
    items: list[TrainingItem],
    family: str,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Default sheet → Run Metadata.
    meta_ws = wb.active
    meta_ws.title = "Run Metadata"
    _write_metadata(meta_ws, family, items, mappings, coverage)

    _write_crosswalk_with_titles(wb.create_sheet("Crosswalk"), mappings, items)
    _write_coverage(wb.create_sheet("Coverage Map"), coverage)
    _write_gaps(wb.create_sheet("Gaps"), gaps)
    _write_surplus(wb.create_sheet("Surplus"), surplus)
    _write_vb_common(wb.create_sheet("Common-V&B Training"), surplus)

    wb.save(out_path)
    return out_path
