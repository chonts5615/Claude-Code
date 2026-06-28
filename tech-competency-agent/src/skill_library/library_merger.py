"""Rolling Master Competency Library merger (spec §30.5).

Promotes a per-run `{run_id}_TechComp_Library_Master.xlsx` (snapshot deliverable)
into the rolling canonical `data/library/TechComp_Library_Master.xlsx`.

Match key: Comp_ID.
- New entry → append.
- Existing entry with material changes → replace row in-place, emit one
  `change_log.jsonl` event per changed field via `src.tracing.change_log`.
- Existing entry not present in this run → leave in place (archival via a
  separate archive workflow; never delete rows here).

Idempotent: re-running an identical run produces zero change_log events.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from src.schemas.competency import TechnicalCompetency
from src.schemas.library import LIBRARY_COLUMNS
from src.tracing.change_log import record_change
from src.utils.branding import alt_row_fill, body_font, header_fill, header_font

logger = logging.getLogger(__name__)

# Fields tracked for change-log emission. Maps schema field -> Excel column name.
_TRACKED_FIELDS: List[Tuple[str, str]] = [
    ("name", "name"),
    ("definition", "definition"),
    ("why_it_matters", "why_it_matters"),
    ("boundary_class", "boundary_class"),
    ("integrity_tag", "integrity_tag"),
]


def _competency_to_row(comp: TechnicalCompetency, run_id: str) -> Dict[str, Any]:
    levels = {pl.level.value: pl for pl in comp.proficiency_levels}
    row: Dict[str, Any] = {col: "" for col in LIBRARY_COLUMNS}
    row.update({
        "competency_id": comp.competency_id,
        "name": comp.name,
        "definition": comp.definition,
        "boundary_class": comp.boundary_class,
        "why_it_matters": comp.why_it_matters,
        "applied_tools": "|".join(comp.applied_scope.tools_methods_tech),
        "applied_standards": "|".join(comp.applied_scope.standards_frameworks),
        "applied_outputs": "|".join(comp.applied_scope.typical_outputs),
        "integrity_tag": comp.integrity_tag,
        "criticality_score": (
            comp.criticality.weighted_score if comp.criticality is not None else ""
        ),
        "last_modified_run": run_id,
    })
    for code in ("L1", "L2", "L3", "L4"):
        if code in levels:
            pl = levels[code]
            row[f"{code}_description"] = pl.description
            row[f"{code}_indicators"] = "|".join(i.text for i in pl.indicators)
    if not row.get("first_published_run"):
        row["first_published_run"] = run_id
    return row


def _load_existing(master_path: Path) -> Dict[str, Dict[str, Any]]:
    if not master_path.exists():
        return {}
    wb = load_workbook(master_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = list(rows[0])
    out: Dict[str, Dict[str, Any]] = {}
    for raw in rows[1:]:
        row = {col: raw[i] if i < len(raw) else "" for i, col in enumerate(header)}
        cid = str(row.get("competency_id") or "").strip()
        if cid:
            out[cid] = row
    return out


def _write_master(master_path: Path, entries: Dict[str, Dict[str, Any]]) -> None:
    master_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "TechComp_Library_Master"
    ws.append(LIBRARY_COLUMNS)
    header_row = ws[1]
    for cell in header_row:
        cell.font = header_font()
        cell.fill = header_fill()
    body = body_font()
    alt = alt_row_fill()
    for idx, cid in enumerate(sorted(entries)):
        row = entries[cid]
        ws.append([row.get(col, "") for col in LIBRARY_COLUMNS])
        excel_row = idx + 2
        for col_idx in range(1, len(LIBRARY_COLUMNS) + 1):
            ws.cell(row=excel_row, column=col_idx).font = body
            if excel_row % 2 == 0:
                ws.cell(row=excel_row, column=col_idx).fill = alt
    wb.save(master_path)


def merge_library_into_master(
    competencies: Iterable[TechnicalCompetency],
    master_path: Path,
    *,
    run_id: str,
    family: Optional[str] = None,
) -> Path:
    """Merge per-run competencies into the rolling Master Competency Library.

    Idempotent: identical inputs → zero change_log events. Emits one
    `change_log.jsonl` event per changed field on update; emits one synthetic
    `record_change(... source="LIBRARY_MERGER", field="<created>")` on add.

    Returns the master path.
    """
    existing = _load_existing(master_path)

    for comp in competencies:
        new_row = _competency_to_row(comp, run_id)
        cid = new_row["competency_id"]
        prior = existing.get(cid)

        if prior is None:
            # Brand-new entry — log a single creation event.
            existing[cid] = new_row
            record_change(
                run_id=run_id,
                competency_id=cid,
                field="<created>",
                before="",
                after=f"{comp.name}: {comp.definition[:80]}",
                source="LIBRARY_MERGER",
                rationale=f"first publication via run {run_id}" + (f" (family {family})" if family else ""),
            )
            continue

        # Update existing row in-place — log each material change.
        # Normalize None → "" so openpyxl's empty-cell behavior doesn't trigger
        # false-positive "changes" on idempotent re-merge.
        def _norm(v: Any) -> str:
            return "" if v is None else str(v).strip()

        any_change = False
        for field, col_name in _TRACKED_FIELDS:
            before = _norm(prior.get(col_name))
            after = _norm(new_row.get(col_name))
            if before != after:
                record_change(
                    run_id=run_id,
                    competency_id=cid,
                    field=col_name,
                    before=before,
                    after=after,
                    source="LIBRARY_MERGER",
                    rationale=f"updated via run {run_id}",
                )
                any_change = True

        # Replace columns; preserve first_published_run.
        first_pub = prior.get("first_published_run") or run_id
        existing[cid] = new_row
        existing[cid]["first_published_run"] = first_pub
        if not any_change:
            # Don't bump last_modified_run if nothing material changed (idempotent).
            existing[cid]["last_modified_run"] = prior.get("last_modified_run") or run_id

    _write_master(master_path, existing)
    return master_path
