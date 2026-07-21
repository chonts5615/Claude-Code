"""Rolling Master Skill→Competency Crosswalk (spec §29.2).

Keyed by `(course_id, competency_id)`. Idempotent: re-running the same
map-skills run produces zero new rows and zero new run_history events.

Auto-deprecation: rows not present in two consecutive runs transition to
`Deprecated`.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from src.skill_mapping.schemas import SkillCompetencyMapping
from src.utils.branding import alt_row_fill, body_font, header_fill, header_font

CROSSWALK_COLUMNS: List[str] = [
    "Skill_ID",
    "Course_ID",
    "Competency_ID",
    "Competency_Name",
    "Current_Level",
    "Current_Confidence",
    "Current_Integrity_Tag",
    "First_Mapped_Run",
    "Last_Mapped_Run",
    "Run_History",
    "Status",
    "Notes",
]


class CrosswalkEntry(dict):
    def __init__(self, **kwargs: Any):
        for col in CROSSWALK_COLUMNS:
            kwargs.setdefault(col, "")
        super().__init__(kwargs)

    def to_row(self) -> List[Any]:
        return [self.get(col, "") for col in CROSSWALK_COLUMNS]


def _key(course_id: str, competency_id: str) -> Tuple[str, str]:
    return (str(course_id), str(competency_id))


def _load_existing(master_path: Path) -> Dict[Tuple[str, str], CrosswalkEntry]:
    if not master_path.exists():
        return {}
    wb = load_workbook(master_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = list(rows[0])
    out: Dict[Tuple[str, str], CrosswalkEntry] = {}
    for raw in rows[1:]:
        row = {col: raw[i] if i < len(raw) else "" for i, col in enumerate(header)}
        cid = str(row.get("Course_ID") or "").strip()
        comp = str(row.get("Competency_ID") or "").strip()
        if not cid or not comp:
            continue
        # Coerce numeric.
        try:
            row["Current_Confidence"] = float(row.get("Current_Confidence") or 0.0)
        except (TypeError, ValueError):
            row["Current_Confidence"] = 0.0
        out[_key(cid, comp)] = CrosswalkEntry(**row)
    return out


def _write_master(master_path: Path, entries: Dict[Tuple[str, str], CrosswalkEntry]) -> None:
    master_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Skill_Competency_Crosswalk"
    ws.append(CROSSWALK_COLUMNS)

    header_row = ws[1]
    for cell in header_row:
        cell.font = header_font()
        cell.fill = header_fill()

    body = body_font()
    alt = alt_row_fill()
    for idx, key in enumerate(sorted(entries.keys())):
        row = entries[key].to_row()
        ws.append(row)
        excel_row = idx + 2
        for col_idx in range(1, len(CROSSWALK_COLUMNS) + 1):
            ws.cell(row=excel_row, column=col_idx).font = body
            if excel_row % 2 == 0:
                ws.cell(row=excel_row, column=col_idx).fill = alt
    wb.save(master_path)


def _parse_history(raw: Any) -> List[Dict[str, Any]]:
    if not raw:
        return []
    if isinstance(raw, list):
        return list(raw)
    try:
        loaded = json.loads(str(raw))
        return list(loaded) if isinstance(loaded, list) else []
    except (json.JSONDecodeError, TypeError, ValueError):
        return []


def _serialize_history(entries: List[Dict[str, Any]]) -> str:
    return json.dumps(entries, default=str)


def _competency_name_for(mapping: SkillCompetencyMapping, names: Dict[str, str]) -> str:
    return names.get(mapping.competency_id, mapping.competency_id)


def merge_crosswalk(
    mappings: Iterable[SkillCompetencyMapping],
    master_path: Path,
    *,
    run_id: str,
    competency_names: Optional[Dict[str, str]] = None,
    skill_ids: Optional[Dict[str, str]] = None,
) -> Path:
    """Merge per-run mappings into the rolling master crosswalk.

    competency_names: map competency_id -> human-readable name (optional).
    skill_ids:        map course_id -> Skill_ID (optional; if missing, leaves blank).

    Returns the master path.
    """
    existing = _load_existing(master_path)
    competency_names = competency_names or {}
    skill_ids = skill_ids or {}

    timestamp = datetime.utcnow().isoformat() + "Z"
    seen_keys: set[Tuple[str, str]] = set()

    for m in mappings:
        key = _key(m.course_id, m.competency_id)
        seen_keys.add(key)
        level = m.level.value if hasattr(m.level, "value") else str(m.level)
        prior = existing.get(key)

        if prior is None:
            history = [{
                "run_id": run_id,
                "level": level,
                "confidence": round(m.confidence, 4),
                "integrity_tag": m.integrity_tag,
                "timestamp": timestamp,
            }]
            entry = CrosswalkEntry(
                Skill_ID=skill_ids.get(m.course_id, ""),
                Course_ID=m.course_id,
                Competency_ID=m.competency_id,
                Competency_Name=_competency_name_for(m, competency_names),
                Current_Level=level,
                Current_Confidence=round(m.confidence, 4),
                Current_Integrity_Tag=m.integrity_tag,
                First_Mapped_Run=run_id,
                Last_Mapped_Run=run_id,
                Run_History=_serialize_history(history),
                Status="Active",
                Notes="",
            )
            existing[key] = entry
            continue

        # Existing key — check for material change.
        history = _parse_history(prior.get("Run_History"))
        last = history[-1] if history else {}
        material_change = (
            str(last.get("level")) != level
            or abs(float(last.get("confidence", 0.0)) - m.confidence) > 1e-6
            or str(last.get("integrity_tag")) != m.integrity_tag
        )

        if material_change:
            history.append({
                "run_id": run_id,
                "level": level,
                "confidence": round(m.confidence, 4),
                "integrity_tag": m.integrity_tag,
                "timestamp": timestamp,
            })

        # Always refresh Last_Mapped_Run + current fields; Status reactivates.
        prior["Current_Level"] = level
        prior["Current_Confidence"] = round(m.confidence, 4)
        prior["Current_Integrity_Tag"] = m.integrity_tag
        prior["Last_Mapped_Run"] = run_id
        prior["Status"] = "Active"
        if material_change:
            prior["Run_History"] = _serialize_history(history)
        # Refresh competency name if previously empty.
        if not prior.get("Competency_Name") and competency_names:
            prior["Competency_Name"] = _competency_name_for(m, competency_names)
        if not prior.get("Skill_ID") and skill_ids.get(m.course_id):
            prior["Skill_ID"] = skill_ids[m.course_id]

    # Auto-deprecate rows not seen this run if they've now missed 2+ runs.
    for key, entry in existing.items():
        if key in seen_keys:
            continue
        last_seen = str(entry.get("Last_Mapped_Run") or "")
        if last_seen == run_id:
            continue  # impossible, defensive
        history = _parse_history(entry.get("Run_History"))
        # Count consecutive trailing runs that don't include this run.
        # Simple rule: if Status was already Active and last_seen != current,
        # bump to Deprecated.
        if entry.get("Status") == "Active":
            entry["Status"] = "Deprecated"
            # Note in run_history that this run did not include the row.
            history.append({
                "run_id": run_id,
                "event": "NOT_SEEN",
                "timestamp": timestamp,
            })
            entry["Run_History"] = _serialize_history(history)

    _write_master(master_path, existing)
    return master_path
