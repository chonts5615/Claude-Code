"""Master Skills Library — parallel canonical library for L&D content (spec §28).

The Master Skills Library is the rolling system of record for every training
item the system has ever observed. It is keyed by `Skill_ID` (stable, opaque)
and additionally carries the vendor `Course_ID` for traceability.

Two physical files:
- `data/library/Skills_Library_Master.xlsx` — rolling canonical Excel.
- `data/library/Skills_Library_Master.jsonl` — append-only event journal.

Merger semantics:
- Match by `Skill_ID`. If absent, derive a deterministic Skill_ID from
  `(Job_Family, Course_ID)` so two runs of the same catalog converge.
- New entry → append row, write JSONL event {"event": "ADD", ...}.
- Existing entry whose material fields changed → replace row in-place, bump
  Version, write JSONL event {"event": "UPDATE", "changed_fields": [...]}.
- Existing entry unchanged → no-op (idempotent — guarantee for spec §21 #25).
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook

from src.skill_mapping.schemas import TrainingItem
from src.utils.branding import alt_row_fill, body_font, header_fill, header_font

SKILLS_LIBRARY_COLUMNS: List[str] = [
    "Skill_ID",
    "Course_ID",
    "Title",
    "Description",
    "Job_Family",
    "Modality",
    "Duration_Hours",
    "Audience_Band",
    "Prerequisites",
    "Vendor",
    "Bloom_Level",
    "Bloom_Confidence",
    "Bloom_Evidence_Verbs",
    "Bloom_Adjustments",
    "Integrity_Tag",
    "First_Seen_Run",
    "Last_Seen_Run",
    "Version",
    "Status",
    "Linked_Competencies",
    "Source_Refs",
    "Notes",
]


# Fields that count as "material" for change detection (Version bump).
_MATERIAL_FIELDS = (
    "Title",
    "Description",
    "Modality",
    "Duration_Hours",
    "Audience_Band",
    "Prerequisites",
    "Vendor",
    "Bloom_Level",
    "Bloom_Confidence",
    "Integrity_Tag",
)


def derive_skill_id(family: str, course_id: str) -> str:
    """Deterministic ID so re-running an identical catalog converges (idempotency)."""
    fam_prefix = re.sub(r"[^A-Za-z]", "", family).upper()[:3] or "GEN"
    digest = hashlib.sha1(f"{family}|{course_id}".encode("utf-8")).hexdigest()[:6]
    return f"SK-{fam_prefix}-{digest.upper()}"


def _norm(v: Any) -> str:
    """Normalize a value for change comparison: treat None and '' as identical."""
    if v is None:
        return ""
    return str(v).strip()


class SkillsLibraryEntry(dict):
    """Lightweight row container — keeps free-form fields without pydantic overhead.

    All values are stored as-is and serialized to the Excel canonical via
    `to_row(columns)`."""

    def __init__(self, **kwargs: Any):
        # Fill missing columns with empty values so the row is always
        # well-shaped against SKILLS_LIBRARY_COLUMNS.
        for col in SKILLS_LIBRARY_COLUMNS:
            kwargs.setdefault(col, "")
        super().__init__(kwargs)

    def to_row(self) -> List[Any]:
        return [self.get(col, "") for col in SKILLS_LIBRARY_COLUMNS]


def _training_item_to_entry(
    item: TrainingItem,
    *,
    family: str,
    run_id: str,
    bloom_level: str = "",
    bloom_confidence: float = 0.0,
    bloom_evidence_verbs: Iterable[str] = (),
    bloom_adjustments: Iterable[str] = (),
    integrity_tag: str = "UNVERIFIABLE",
    linked_competencies: Iterable[str] = (),
) -> SkillsLibraryEntry:
    return SkillsLibraryEntry(
        Skill_ID=derive_skill_id(family, item.course_id),
        Course_ID=item.course_id,
        Title=item.title,
        Description=item.description or "",
        Job_Family=family,
        Modality=item.modality,
        Duration_Hours=float(item.duration_hours or 0.0),
        Audience_Band=item.audience_band or "",
        Prerequisites="|".join(item.prerequisites or []),
        Vendor=item.vendor or "",
        Bloom_Level=bloom_level,
        Bloom_Confidence=round(float(bloom_confidence), 4),
        Bloom_Evidence_Verbs="|".join(bloom_evidence_verbs),
        Bloom_Adjustments="|".join(bloom_adjustments),
        Integrity_Tag=integrity_tag,
        First_Seen_Run=run_id,
        Last_Seen_Run=run_id,
        Version=1,
        Status="Active",
        Linked_Competencies="|".join(linked_competencies),
        Source_Refs="",
        Notes="",
    )


def _load_existing(master_path: Path) -> Dict[str, SkillsLibraryEntry]:
    if not master_path.exists():
        return {}
    wb = load_workbook(master_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = list(rows[0])
    out: Dict[str, SkillsLibraryEntry] = {}
    for raw in rows[1:]:
        row_dict = {col: raw[i] if i < len(raw) else "" for i, col in enumerate(header)}
        sid = str(row_dict.get("Skill_ID") or "").strip()
        if not sid:
            continue
        # Coerce numeric columns.
        for k in ("Duration_Hours", "Bloom_Confidence"):
            try:
                row_dict[k] = float(row_dict.get(k) or 0.0)
            except (TypeError, ValueError):
                row_dict[k] = 0.0
        try:
            row_dict["Version"] = int(row_dict.get("Version") or 1)
        except (TypeError, ValueError):
            row_dict["Version"] = 1
        out[sid] = SkillsLibraryEntry(**row_dict)
    return out


def _write_master(master_path: Path, entries: Dict[str, SkillsLibraryEntry]) -> None:
    master_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills_Library"
    ws.append(SKILLS_LIBRARY_COLUMNS)

    header_row = ws[1]
    for cell in header_row:
        cell.font = header_font()
        cell.fill = header_fill()

    body = body_font()
    alt = alt_row_fill()
    for idx, sid in enumerate(sorted(entries)):
        row = entries[sid].to_row()
        ws.append(row)
        excel_row = idx + 2
        for col_idx in range(1, len(SKILLS_LIBRARY_COLUMNS) + 1):
            ws.cell(row=excel_row, column=col_idx).font = body
            if excel_row % 2 == 0:
                ws.cell(row=excel_row, column=col_idx).fill = alt
    wb.save(master_path)


def _journal_path(master_path: Path) -> Path:
    return master_path.with_suffix(".jsonl")


def _emit_journal(master_path: Path, event: Dict[str, Any]) -> None:
    event.setdefault("at_utc", datetime.utcnow().isoformat() + "Z")
    jpath = _journal_path(master_path)
    jpath.parent.mkdir(parents=True, exist_ok=True)
    with jpath.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(event, default=str) + "\n")


def merge_skills_into_master(
    items: Iterable[TrainingItem],
    master_path: Path,
    *,
    run_id: str,
    family: str,
    bloom_estimates: Optional[Dict[str, Any]] = None,
    linked_by_course: Optional[Dict[str, List[str]]] = None,
) -> Path:
    """Merge a batch of TrainingItems into the rolling Master Skills Library.

    bloom_estimates: map course_id -> BloomLevelEstimate (optional; if present,
        fills Bloom_* columns and integrity tag derives from bloom confidence).
    linked_by_course: map course_id -> [competency_id, ...] from the current
        crosswalk run (optional; refreshes Linked_Competencies).

    Returns the master path. Emits one JSONL event per ADD/UPDATE.
    """
    existing = _load_existing(master_path)
    bloom_estimates = bloom_estimates or {}
    linked_by_course = linked_by_course or {}
    add_count = 0
    update_count = 0

    for item in items:
        be = bloom_estimates.get(item.course_id)
        bloom_level = getattr(be, "level", None)
        bloom_level_str = bloom_level.value if hasattr(bloom_level, "value") else str(bloom_level or "")
        bloom_conf = float(getattr(be, "confidence", 0.0) or 0.0)
        evidence = list(getattr(be, "evidence_verbs", []) or [])
        adjustments = list(getattr(be, "adjustments_applied", []) or [])

        # Integrity tag from bloom confidence band.
        if bloom_conf >= 0.7:
            tag = "CONFIRMED"
        elif bloom_conf >= 0.55:
            tag = "UNVERIFIABLE"
        else:
            tag = "FLAGGED" if be is not None else "UNVERIFIABLE"

        new_entry = _training_item_to_entry(
            item,
            family=family,
            run_id=run_id,
            bloom_level=bloom_level_str,
            bloom_confidence=bloom_conf,
            bloom_evidence_verbs=evidence,
            bloom_adjustments=adjustments,
            integrity_tag=tag,
            linked_competencies=linked_by_course.get(item.course_id, []),
        )
        sid = new_entry["Skill_ID"]
        prior = existing.get(sid)
        if prior is None:
            existing[sid] = new_entry
            add_count += 1
            _emit_journal(master_path, {
                "event": "ADD",
                "run_id": run_id,
                "skill_id": sid,
                "course_id": item.course_id,
                "title": item.title,
            })
            continue

        # Compare material fields.
        changed = []
        for field in _MATERIAL_FIELDS:
            prior_val = prior.get(field, "")
            new_val = new_entry.get(field, "")
            # Normalize for comparison — openpyxl returns None for empty cells,
            # but new entries use "" or 0.0 as sentinels.
            if isinstance(prior_val, float) or isinstance(new_val, float):
                if abs(float(prior_val or 0.0) - float(new_val or 0.0)) > 1e-6:
                    changed.append(field)
            else:
                if _norm(prior_val) != _norm(new_val):
                    changed.append(field)

        # Always refresh Last_Seen_Run and Linked_Competencies (no version bump).
        prior["Last_Seen_Run"] = run_id
        prior["Linked_Competencies"] = new_entry["Linked_Competencies"]

        if changed:
            for field in changed:
                prior[field] = new_entry[field]
            try:
                prior["Version"] = int(prior.get("Version") or 1) + 1
            except (TypeError, ValueError):
                prior["Version"] = 2
            update_count += 1
            _emit_journal(master_path, {
                "event": "UPDATE",
                "run_id": run_id,
                "skill_id": sid,
                "changed_fields": changed,
                "new_version": prior["Version"],
            })

    _write_master(master_path, existing)
    _emit_journal(master_path, {
        "event": "MERGE_SUMMARY",
        "run_id": run_id,
        "adds": add_count,
        "updates": update_count,
        "total": len(existing),
    })
    return master_path
