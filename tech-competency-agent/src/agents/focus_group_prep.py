"""Phase 6G: Focus Group Prep Agent (v3.1)."""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from src.agents.base import BaseAgent
from src.schemas.feedback import FeedbackBatch, FeedbackItem
from src.schemas.run_state import RunState
from src.utils.branding import alt_row_fill, body_font, header_fill, header_font


class FocusGroupPrepAgent(BaseAgent):
    """Phase 6G — Build a focus group package from DISCUSS / deferred feedback."""

    def __init__(self, agent_id: str = "phase6G_focus_group_prep",
                 step_name: str = "Phase 6G — Focus Group Prep"):
        super().__init__(agent_id, step_name)

    def execute(self, state: RunState) -> RunState:
        state.current_step = self.agent_id

        batch = self._load_feedback_batch(state)
        if batch is None:
            self.add_flag(
                state,
                severity="WARNING",
                flag_type="FEEDBACK_BATCH_MISSING",
                message="No FeedbackBatch artifact present; skipping focus group prep.",
            )
            try:
                setattr(state.artifacts, "focus_group_package", None)
            except Exception:
                pass
            return state

        discuss_items = [
            item for item in batch.items
            if item.disposition == "DISCUSS" or self._is_deferred(item)
        ]
        if not discuss_items:
            try:
                setattr(state.artifacts, "focus_group_package", None)
            except Exception:
                pass
            return state

        topics = self._aggregate_by_topic(discuss_items)
        attendees = self._recommend_attendees(batch.family, topics)

        # TODO(v3.1): integrate LLM judgment to compose stronger discussion
        # questions and disambiguate topic clustering.
        questions = self._draft_questions(topics)
        evidence = self._collect_evidence(topics)

        output_path = Path(f"data/output/{state.run_id}_6G_focus_group_package.xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._write_workbook(output_path, batch, topics, questions, evidence, attendees)

        try:
            setattr(state.artifacts, "focus_group_package", output_path)
        except Exception:
            pass

        return state

    @staticmethod
    def _is_deferred(item: FeedbackItem) -> bool:
        rationale = (item.rationale or "").lower()
        return "defer" in rationale or "deferred" in rationale

    @staticmethod
    def _aggregate_by_topic(items: List[FeedbackItem]) -> Dict[str, List[FeedbackItem]]:
        topics: Dict[str, List[FeedbackItem]] = defaultdict(list)
        for item in items:
            topic = item.target_competency_id or item.target_field or "GENERAL"
            topics[str(topic)].append(item)
        return topics

    @staticmethod
    def _recommend_attendees(family: str, topics: Dict[str, List[FeedbackItem]]) -> List[str]:
        attendees = {f"{family} SME Lead", f"{family} Hiring Manager"}
        for items in topics.values():
            for item in items:
                if item.sme_role:
                    attendees.add(item.sme_role)
                if item.is_anchor_sme and item.sme_name:
                    attendees.add(f"Anchor SME: {item.sme_name}")
        return sorted(attendees)

    @staticmethod
    def _draft_questions(topics: Dict[str, List[FeedbackItem]]) -> List[Dict[str, str]]:
        questions: List[Dict[str, str]] = []
        for topic, items in topics.items():
            questions.append({
                "topic": topic,
                "question": (
                    f"For {topic}: SMEs flagged this for discussion — what is the right "
                    "scope, level, and language for this competency?"
                ),
                "open_items": str(len(items)),
            })
        return questions

    @staticmethod
    def _collect_evidence(topics: Dict[str, List[FeedbackItem]]) -> List[Dict[str, str]]:
        evidence: List[Dict[str, str]] = []
        for topic, items in topics.items():
            for item in items:
                evidence.append({
                    "topic": topic,
                    "feedback_id": item.feedback_id,
                    "sme": item.sme_name or "(anonymous)",
                    "role": item.sme_role or "",
                    "verbatim": item.verbatim_comment,
                    "proposed": item.proposed_text or "",
                })
        return evidence

    @staticmethod
    def _write_workbook(
        path: Path,
        batch: FeedbackBatch,
        topics: Dict[str, List[FeedbackItem]],
        questions: List[Dict[str, str]],
        evidence: List[Dict[str, str]],
        attendees: List[str],
    ) -> None:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("openpyxl is required to write the focus group package") from exc

        wb = openpyxl.Workbook()
        cover = wb.active
        cover.title = "Cover"
        cover_rows = [
            ["Focus Group Package"],
            ["Run ID", batch.run_id],
            ["Family", batch.family],
            ["Stage", batch.stage],
            ["Topics", str(len(topics))],
            ["Generated UTC", datetime.utcnow().isoformat()],
        ]
        for r_idx, row in enumerate(cover_rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = cover.cell(row=r_idx, column=c_idx, value=value)
                cell.font = header_font() if r_idx == 1 else body_font()
                if r_idx == 1:
                    cell.fill = header_fill()

        FocusGroupPrepAgent._write_sheet(
            wb, "Questions",
            ["topic", "question", "open_items"],
            [[q["topic"], q["question"], q["open_items"]] for q in questions],
        )
        FocusGroupPrepAgent._write_sheet(
            wb, "Evidence",
            ["topic", "feedback_id", "sme", "role", "verbatim", "proposed"],
            [[e["topic"], e["feedback_id"], e["sme"], e["role"], e["verbatim"], e["proposed"]]
             for e in evidence],
        )
        FocusGroupPrepAgent._write_sheet(
            wb, "Attendees",
            ["recommended_attendee"],
            [[a] for a in attendees],
        )
        wb.save(path)

    @staticmethod
    def _write_sheet(wb, title: str, headers: List[str], rows: List[List[str]]) -> None:
        sheet = wb.create_sheet(title)
        for c_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=c_idx, value=header)
            cell.fill = header_fill()
            cell.font = header_font()
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                cell.font = body_font()
                if r_idx % 2 == 0:
                    cell.fill = alt_row_fill()

    @staticmethod
    def _load_feedback_batch(state: RunState) -> FeedbackBatch | None:
        feedback_artifact = getattr(state.artifacts, "feedback_batch", None)
        if not feedback_artifact:
            return None
        path = Path(str(feedback_artifact))
        if not path.exists():
            return None
        try:
            with open(path) as fh:
                return FeedbackBatch.model_validate(json.load(fh))
        except Exception:
            return None

    def get_system_prompt(self) -> str:
        return (
            "You are the Focus Group Prep Operator for v3.1 Phase 6G. You only run when the "
            "FeedbackBatch contains DISCUSS items or unresolved deferrals — your job is to turn "
            "those conversational threads into a structured agenda for an SME working session.\n\n"
            "Steps:\n"
            "1. Load the FeedbackBatch and filter for DISCUSS or deferred items.\n"
            "2. Aggregate the items by topic (target_competency_id, then target_field).\n"
            "3. Draft a discussion question per topic and gather verbatim evidence excerpts.\n"
            "4. Recommend attendees by role family, including anchor SMEs by name.\n"
            "5. Write a Cargill-branded Excel package and update state.artifacts.focus_group_package.\n\n"
            "Quality standards: never paraphrase verbatim comments; always include the feedback_id "
            "for traceability; if no DISCUSS items exist, leave the artifact pointer null."
        )
