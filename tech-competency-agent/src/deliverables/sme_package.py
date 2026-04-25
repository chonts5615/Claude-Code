"""SME Review Package writer.

Mirrors the family package layout but adds editable feedback columns
(``disposition``, ``comment``, ``proposed_text``) so SMEs can mark up
competencies in place. The resulting workbook is fed back into
:class:`src.schemas.feedback.FeedbackBatch` after the review window closes.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schemas.competency import TechnicalCompetency
from src.utils.branding import (
    alt_row_fill,
    body_font,
    header_fill,
    header_font,
)


_OVERVIEW_HEADERS = ["Field", "Value"]
_REVIEW_HEADERS = [
    "competency_id",
    "name",
    "boundary_class",
    "definition",
    "L1_indicators",
    "L2_indicators",
    "L3_indicators",
    "L4_indicators",
    # Editable feedback columns:
    "disposition",
    "comment",
    "proposed_text",
]
_LEVEL_DETAIL_HEADERS = [
    "competency_id",
    "level",
    "description",
    "indicators",
    "disposition",
    "comment",
    "proposed_text",
]


def _apply_header_row(ws: Worksheet, n_cols: int) -> None:
    h_fill: PatternFill = header_fill()
    h_font: Font = header_font()
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26


def _apply_alt_rows(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    alt_fill: PatternFill = alt_row_fill()
    b_font: Font = body_font()
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = b_font
            cell.alignment = wrap
            if row % 2 == 0:
                cell.fill = alt_fill


def _set_widths(ws: Worksheet, widths: List[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _pipe(items) -> str:
    return " | ".join(items)


def _level_indicators(comp: TechnicalCompetency, code: str) -> str:
    for lvl in comp.proficiency_levels:
        if lvl.level.value == code:
            return _pipe(ind.text for ind in lvl.indicators)
    return ""


def _build_overview(wb: Workbook, family: str, run_id: str, n_competencies: int) -> None:
    ws = wb.create_sheet("Overview")
    ws.append(_OVERVIEW_HEADERS)
    rows = [
        ("Family", family),
        ("Run ID", run_id),
        ("Stage", "SME Review"),
        ("Competency count", n_competencies),
        ("Instructions",
         "Fill the disposition column with KEEP / EDIT / GAP / DISCUSS / REJECT. "
         "Add a comment and (when EDIT) the proposed_text. Do not modify other columns."),
    ]
    for row in rows:
        ws.append(list(row))
    _apply_header_row(ws, n_cols=2)
    _apply_alt_rows(ws, n_rows=len(rows), n_cols=2)
    _set_widths(ws, [24, 90])
    ws.freeze_panes = "A2"


def _build_review(wb: Workbook, competencies: list[TechnicalCompetency]) -> None:
    ws = wb.create_sheet("Competency Review")
    ws.append(_REVIEW_HEADERS)
    for comp in competencies:
        ws.append([
            comp.competency_id,
            comp.name,
            comp.boundary_class,
            comp.definition,
            _level_indicators(comp, "L1"),
            _level_indicators(comp, "L2"),
            _level_indicators(comp, "L3"),
            _level_indicators(comp, "L4"),
            "",  # disposition
            "",  # comment
            "",  # proposed_text
        ])
    _apply_header_row(ws, n_cols=len(_REVIEW_HEADERS))
    _apply_alt_rows(ws, n_rows=len(competencies), n_cols=len(_REVIEW_HEADERS))
    _set_widths(ws, [18, 24, 14, 60, 60, 60, 60, 60, 14, 40, 40])
    ws.freeze_panes = "C2"


def _build_level_detail(wb: Workbook, competencies: list[TechnicalCompetency]) -> None:
    ws = wb.create_sheet("Level Detail")
    ws.append(_LEVEL_DETAIL_HEADERS)
    n_rows = 0
    for comp in competencies:
        for lvl in comp.proficiency_levels:
            ws.append([
                comp.competency_id,
                lvl.level.value,
                lvl.description,
                _pipe(ind.text for ind in lvl.indicators),
                "",
                "",
                "",
            ])
            n_rows += 1
    _apply_header_row(ws, n_cols=len(_LEVEL_DETAIL_HEADERS))
    _apply_alt_rows(ws, n_rows=n_rows, n_cols=len(_LEVEL_DETAIL_HEADERS))
    _set_widths(ws, [18, 8, 50, 60, 14, 40, 40])
    ws.freeze_panes = "C2"


def write_sme_package(
    out_dir: Path,
    family: str,
    competencies: list[TechnicalCompetency],
    run_id: str,
) -> dict[str, Path]:
    """Write the SME review package to ``out_dir``.

    Args:
        out_dir: Directory for the package (created if missing).
        family: Job family label.
        competencies: Competencies under review.
        run_id: Run identifier.

    Returns:
        Mapping of artifact name to absolute path.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    pkg_path = out_dir / f"{family}_SME_Review_Package.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    _build_overview(wb, family=family, run_id=run_id, n_competencies=len(competencies))
    _build_review(wb, competencies=competencies)
    _build_level_detail(wb, competencies=competencies)

    wb.save(pkg_path)
    return {"sme_review_package": pkg_path}
