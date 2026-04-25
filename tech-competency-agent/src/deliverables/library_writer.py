"""Master Library writer — produces TechComp_Library_Master.xlsx.

Writes the 23-column canonical library defined in
:data:`src.schemas.library.LIBRARY_COLUMNS`, one row per
:class:`TechnicalCompetency`.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schemas.competency import TechnicalCompetency
from src.schemas.library import LIBRARY_COLUMNS
from src.utils.branding import (
    alt_row_fill,
    body_font,
    header_fill,
    header_font,
)

SHEET_NAME = "Library Master"

# Reasonable column widths keyed by canonical field name.
_COLUMN_WIDTHS: dict[str, int] = {
    "competency_id": 18,
    "name": 20,
    "family": 16,
    "boundary_class": 14,
    "definition": 60,
    "why_it_matters": 50,
    "L1_description": 40,
    "L1_indicators": 80,
    "L2_description": 40,
    "L2_indicators": 80,
    "L3_description": 40,
    "L3_indicators": 80,
    "L4_description": 40,
    "L4_indicators": 80,
    "applied_tools": 30,
    "applied_standards": 30,
    "applied_outputs": 30,
    "criticality_score": 14,
    "integrity_tag": 14,
    "source_refs": 30,
    "rosetta_aliases": 30,
    "first_published_run": 18,
    "last_modified_run": 18,
}


def _pipe(items: Iterable[str]) -> str:
    return " | ".join(s for s in items if s)


def _level_lookup(comp: TechnicalCompetency, code: str) -> tuple[str, str]:
    """Return (description, pipe-delimited indicators) for the given level code."""
    for lvl in comp.proficiency_levels:
        if lvl.level.value == code:
            indicators = _pipe(ind.text for ind in lvl.indicators)
            return lvl.description, indicators
    return "", ""


def _competency_row(comp: TechnicalCompetency, family: str, run_id: str) -> List[object]:
    l1_desc, l1_ind = _level_lookup(comp, "L1")
    l2_desc, l2_ind = _level_lookup(comp, "L2")
    l3_desc, l3_ind = _level_lookup(comp, "L3")
    l4_desc, l4_ind = _level_lookup(comp, "L4")

    crit_score: float | str = ""
    if comp.criticality is not None:
        crit_score = round(comp.criticality.weighted_score, 4)

    source_refs = _pipe(s.source_id for s in []) if False else ""
    # source_refs is populated downstream from CompetencyLibraryEntry; technical
    # competencies don't carry SourceEvidence directly in v3.1 schema.

    return [
        comp.competency_id,
        comp.name,
        family,
        comp.boundary_class,
        comp.definition,
        comp.why_it_matters,
        l1_desc,
        l1_ind,
        l2_desc,
        l2_ind,
        l3_desc,
        l3_ind,
        l4_desc,
        l4_ind,
        _pipe(comp.applied_scope.tools_methods_tech),
        _pipe(comp.applied_scope.standards_frameworks),
        _pipe(comp.applied_scope.typical_outputs),
        crit_score,
        comp.integrity_tag,
        source_refs,
        "",  # rosetta_aliases — populated by rosetta_stone writer downstream
        run_id,
        run_id,
    ]


def _apply_branding(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    h_fill: PatternFill = header_fill()
    h_font: Font = header_font()
    alt_fill: PatternFill = alt_row_fill()
    b_font: Font = body_font()
    wrap = Alignment(wrap_text=True, vertical="top")

    # Header row
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Body rows
    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = b_font
            cell.alignment = wrap
            if row % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    for idx, field in enumerate(LIBRARY_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _COLUMN_WIDTHS.get(field, 20)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"


def write_library(
    out_path: Path,
    competencies: list[TechnicalCompetency],
    family: str,
    run_id: str,
) -> Path:
    """Write the 23-column TechComp_Library_Master.xlsx.

    Args:
        out_path: Destination file path (parents are created if missing).
        competencies: Technical competencies to emit (one row each).
        family: Job family label written into column ``family``.
        run_id: Run identifier written into ``first_published_run`` and
            ``last_modified_run``.

    Returns:
        The resolved output path.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(LIBRARY_COLUMNS)
    for comp in competencies:
        ws.append(_competency_row(comp, family=family, run_id=run_id))

    _apply_branding(ws, n_rows=len(competencies), n_cols=len(LIBRARY_COLUMNS))

    wb.save(out_path)
    return out_path
