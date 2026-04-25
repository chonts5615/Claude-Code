"""Job Family Package writer.

Produces ``{family}_Job_Family_Package.xlsx`` with four sheets:

* ``Overview`` — family-level header, run id, counts.
* ``Jobs`` — one row per job in scope.
* ``Competencies`` — one row per technical competency (subset of library cols).
* ``EF Coverage Map`` — job x competency matrix marking primary/secondary/supporting.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schemas.competency import TechnicalCompetency
from src.utils.branding import (
    alt_row_fill,
    body_font,
    header_fill,
    header_font,
)


_OVERVIEW_HEADERS = ["Field", "Value"]
_JOBS_HEADERS = [
    "job_id",
    "job_title",
    "job_family",
    "sub_family",
    "level",
    "ef_count",
    "source_doc",
]
_COMPETENCIES_HEADERS = [
    "competency_id",
    "name",
    "boundary_class",
    "definition",
    "why_it_matters",
    "criticality_score",
    "integrity_tag",
]


def _apply_header_row(ws: Worksheet, n_cols: int) -> None:
    h_fill: PatternFill = header_fill()
    h_font: Font = header_font()
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26


def _apply_alt_rows(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    alt_fill: PatternFill = alt_row_fill()
    b_font: Font = body_font()
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = b_font
            cell.alignment = wrap
            if row % 2 == 0:
                cell.fill = alt_fill


def _set_widths(ws: Worksheet, widths: List[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build_overview(
    wb: Workbook,
    family: str,
    jobs: list[dict],
    competencies: list[TechnicalCompetency],
    run_id: str,
) -> None:
    ws = wb.create_sheet("Overview")
    ws.append(_OVERVIEW_HEADERS)
    rows = [
        ("Family", family),
        ("Run ID", run_id),
        ("Job count", len(jobs)),
        ("Competency count", len(competencies)),
        ("Schema version", "v3.1"),
    ]
    for row in rows:
        ws.append(list(row))
    _apply_header_row(ws, n_cols=2)
    _apply_alt_rows(ws, n_rows=len(rows), n_cols=2)
    _set_widths(ws, [24, 60])
    ws.freeze_panes = "A2"


def _build_jobs(wb: Workbook, jobs: list[dict]) -> None:
    ws = wb.create_sheet("Jobs")
    ws.append(_JOBS_HEADERS)
    for job in jobs:
        ws.append([job.get(h, "") for h in _JOBS_HEADERS])
    _apply_header_row(ws, n_cols=len(_JOBS_HEADERS))
    _apply_alt_rows(ws, n_rows=len(jobs), n_cols=len(_JOBS_HEADERS))
    _set_widths(ws, [16, 36, 18, 18, 10, 10, 30])
    ws.freeze_panes = "C2"


def _build_competencies(wb: Workbook, competencies: list[TechnicalCompetency]) -> None:
    ws = wb.create_sheet("Competencies")
    ws.append(_COMPETENCIES_HEADERS)
    for comp in competencies:
        crit = ""
        if comp.criticality is not None:
            crit = round(comp.criticality.weighted_score, 4)
        ws.append([
            comp.competency_id,
            comp.name,
            comp.boundary_class,
            comp.definition,
            comp.why_it_matters,
            crit,
            comp.integrity_tag,
        ])
    _apply_header_row(ws, n_cols=len(_COMPETENCIES_HEADERS))
    _apply_alt_rows(ws, n_rows=len(competencies), n_cols=len(_COMPETENCIES_HEADERS))
    _set_widths(ws, [18, 24, 14, 60, 50, 14, 14])
    ws.freeze_panes = "C2"


def _build_coverage_map(
    wb: Workbook,
    jobs: list[dict],
    competencies: list[TechnicalCompetency],
) -> None:
    ws = wb.create_sheet("EF Coverage Map")
    headers = ["competency_id", "name"] + [j.get("job_id", "") for j in jobs]
    ws.append(headers)

    # Pre-index responsibility traces by (competency_id, job_id) ~= responsibility_id.
    # Convention: responsibility_id encodes the job_id prefix.
    job_ids = [j.get("job_id", "") for j in jobs]
    for comp in competencies:
        contributions: Dict[str, str] = {}
        for trace in comp.responsibility_trace:
            for jid in job_ids:
                if trace.responsibility_id.startswith(jid):
                    contributions[jid] = trace.contribution
        row = [comp.competency_id, comp.name] + [contributions.get(jid, "") for jid in job_ids]
        ws.append(row)

    _apply_header_row(ws, n_cols=len(headers))
    _apply_alt_rows(ws, n_rows=len(competencies), n_cols=len(headers))
    _set_widths(ws, [18, 24] + [14] * len(jobs))
    ws.freeze_panes = "C2"


def write_family_package(
    out_dir: Path,
    family: str,
    jobs: list[dict],
    competencies: list[TechnicalCompetency],
    run_id: str,
) -> dict[str, Path]:
    """Write the per-family deliverable bundle.

    Args:
        out_dir: Directory to write into (created if missing).
        family: Job family name (used in filename).
        jobs: List of job dicts; expected keys per :data:`_JOBS_HEADERS`.
        competencies: Technical competencies belonging to the family.
        run_id: Run identifier.

    Returns:
        Mapping of artifact name to absolute path.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    pkg_path = out_dir / f"{family}_Job_Family_Package.xlsx"

    wb = Workbook()
    # Replace default empty sheet
    wb.remove(wb.active)

    _build_overview(wb, family=family, jobs=jobs, competencies=competencies, run_id=run_id)
    _build_jobs(wb, jobs=jobs)
    _build_competencies(wb, competencies=competencies)
    _build_coverage_map(wb, jobs=jobs, competencies=competencies)

    wb.save(pkg_path)
    return {"job_family_package": pkg_path}
