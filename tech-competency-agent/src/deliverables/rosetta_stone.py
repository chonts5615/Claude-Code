"""Rosetta Stone writer — cross-family naming map.

Each row records that ``alias`` (used in ``family``) maps to ``canonical_name``
in the master library. ``definition_match_score`` is the cosine similarity
between the alias' definition and the canonical's definition; ``action`` is
one of ``REUSE``, ``DIFFERENTIATE``, ``RENAME``.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.branding import (
    alt_row_fill,
    body_font,
    header_fill,
    header_font,
)

SHEET_NAME = "Rosetta Stone"

ROSETTA_COLUMNS: List[str] = [
    "canonical_name",
    "family",
    "alias",
    "definition_match_score",
    "action",
]

_WIDTHS: List[int] = [28, 18, 28, 18, 18]
_VALID_ACTIONS = {"REUSE", "DIFFERENTIATE", "RENAME"}


def _apply_branding(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    h_fill: PatternFill = header_fill()
    h_font: Font = header_font()
    alt_fill: PatternFill = alt_row_fill()
    b_font: Font = body_font()
    wrap = Alignment(wrap_text=True, vertical="top")

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26

    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = b_font
            cell.alignment = wrap
            if row % 2 == 0:
                cell.fill = alt_fill

    for idx, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "B2"


def write_rosetta(out_path: Path, mappings: list[dict]) -> Path:
    """Write the Rosetta Stone cross-family naming map.

    Args:
        out_path: Output xlsx path (parents created if missing).
        mappings: List of dicts with keys from :data:`ROSETTA_COLUMNS`.
            ``action`` must be one of REUSE / DIFFERENTIATE / RENAME.

    Returns:
        The resolved output path.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(ROSETTA_COLUMNS)
    for m in mappings:
        action = m.get("action", "")
        if action and action not in _VALID_ACTIONS:
            raise ValueError(
                f"Rosetta action must be one of {sorted(_VALID_ACTIONS)}; got {action!r}"
            )
        ws.append([m.get(col, "") for col in ROSETTA_COLUMNS])

    _apply_branding(ws, n_rows=len(mappings), n_cols=len(ROSETTA_COLUMNS))

    wb.save(out_path)
    return out_path
