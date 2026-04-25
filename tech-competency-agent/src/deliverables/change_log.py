"""Change Log writer.

Every edit applied between SME R1, R2 and FINAL gets one row here. The
caller is responsible for assigning ``change_id`` (monotonic) and supplying
``before_text`` / ``after_text`` exactly as they appear in the library.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.branding import (
    alt_row_fill,
    body_font,
    header_fill,
    header_font,
)

SHEET_NAME = "Change Log"

CHANGE_LOG_COLUMNS: List[str] = [
    "change_id",
    "run_id",
    "timestamp_utc",
    "source_sme",
    "target_competency_id",
    "target_field",
    "before_text",
    "after_text",
    "rationale",
    "disposition",
]

_WIDTHS: List[int] = [12, 16, 22, 22, 20, 18, 60, 60, 50, 14]


def _apply_branding(ws: Worksheet, n_rows: int, n_cols: int) -> None:
    h_fill: PatternFill = header_fill()
    h_font: Font = header_font()
    alt_fill: PatternFill = alt_row_fill()
    b_font: Font = body_font()
    wrap = Alignment(wrap_text=True, vertical="top")

    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26

    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = b_font
            cell.alignment = wrap
            if row % 2 == 0:
                cell.fill = alt_fill

    for idx, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "C2"


def write_change_log(out_path: Path, changes: list[dict], run_id: str) -> Path:
    """Write the audit-grade change log.

    Args:
        out_path: Output xlsx path (parents created if missing).
        changes: List of change dicts; missing keys default to empty string.
            Each change should provide keys from :data:`CHANGE_LOG_COLUMNS`.
        run_id: Run identifier; written into any row missing ``run_id``.

    Returns:
        The resolved output path.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(CHANGE_LOG_COLUMNS)
    for change in changes:
        row = []
        for col in CHANGE_LOG_COLUMNS:
            if col == "run_id":
                row.append(change.get("run_id", run_id))
            else:
                row.append(change.get(col, ""))
        ws.append(row)

    _apply_branding(ws, n_rows=len(changes), n_cols=len(CHANGE_LOG_COLUMNS))

    wb.save(out_path)
    return out_path
