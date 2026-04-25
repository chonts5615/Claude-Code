"""BCO Ledger writer — emits the Boundary / Coverage / Overlap workbook."""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schemas.bco_ledger import BCOLedger
from src.utils.branding import (
    alt_row_fill,
    body_font,
    header_fill,
    header_font,
)


_BOUNDARY_HEADERS = [
    "competency_id",
    "competency_name",
    "classification",
    "confidence",
    "rationale",
]
_BOUNDARY_WIDTHS = [18, 28, 16, 12, 60]

_COVERAGE_HEADERS = [
    "job_id",
    "job_title",
    "family",
    "technical_ef_count",
    "technical_ef_covered",
    "coverage_rate",
    "uncovered_ef_ids",
    "meets_90_threshold",
]
_COVERAGE_WIDTHS = [16, 32, 18, 14, 14, 12, 36, 16]

_OVERLAP_HEADERS = [
    "competency_id_a",
    "competency_id_b",
    "similarity_score",
    "severity",
    "resolution",
]
_OVERLAP_WIDTHS = [18, 18, 14, 12, 50]


def _apply_branding(
    ws: Worksheet, n_rows: int, widths: List[int], freeze: str = "B2"
) -> None:
    h_fill: PatternFill = header_fill()
    h_font: Font = header_font()
    alt_fill: PatternFill = alt_row_fill()
    b_font: Font = body_font()
    wrap = Alignment(wrap_text=True, vertical="top")

    n_cols = len(widths)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26

    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = b_font
            cell.alignment = wrap
            if row % 2 == 0:
                cell.fill = alt_fill

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = freeze


def _write_boundary(wb: Workbook, ledger: BCOLedger) -> None:
    ws = wb.create_sheet("Boundary")
    ws.append(_BOUNDARY_HEADERS)
    for entry in ledger.boundary:
        ws.append([
            entry.competency_id,
            entry.competency_name,
            entry.classification,
            round(entry.confidence, 4),
            entry.rationale,
        ])
    _apply_branding(ws, n_rows=len(ledger.boundary), widths=_BOUNDARY_WIDTHS)


def _write_coverage(wb: Workbook, ledger: BCOLedger) -> None:
    ws = wb.create_sheet("Coverage")
    ws.append(_COVERAGE_HEADERS)
    for entry in ledger.coverage:
        ws.append([
            entry.job_id,
            entry.job_title,
            entry.family,
            entry.technical_ef_count,
            entry.technical_ef_covered,
            round(entry.coverage_rate, 4),
            " | ".join(entry.uncovered_ef_ids),
            entry.meets_90_threshold,
        ])
    _apply_branding(ws, n_rows=len(ledger.coverage), widths=_COVERAGE_WIDTHS)


def _write_overlap(wb: Workbook, ledger: BCOLedger) -> None:
    ws = wb.create_sheet("Overlap")
    ws.append(_OVERLAP_HEADERS)
    for entry in ledger.overlap:
        ws.append([
            entry.competency_id_a,
            entry.competency_id_b,
            round(entry.similarity_score, 4),
            entry.severity,
            entry.resolution or "",
        ])
    _apply_branding(ws, n_rows=len(ledger.overlap), widths=_OVERLAP_WIDTHS)


def write_bco_ledger(out_path: Path, ledger: BCOLedger) -> Path:
    """Write the BCO Ledger to ``out_path`` (xlsx, three sheets).

    Args:
        out_path: Output file path (parents created if missing).
        ledger: Populated :class:`BCOLedger`.

    Returns:
        The resolved output path.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _write_boundary(wb, ledger)
    _write_coverage(wb, ledger)
    _write_overlap(wb, ledger)

    wb.save(out_path)
    return out_path
